from fastapi import FastAPI, APIRouter, HTTPException, Header, Response, Request
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from typing import List, Optional, Dict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from io import StringIO, BytesIO
import csv
import json
import os
import logging
import uuid
import bcrypt
import jwt
import httpx
import math
import calendar

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app):
    # Startup
    yield
    # Shutdown
    client.close()

# Create the main app
app = FastAPI(lifespan=lifespan)
api_router = APIRouter(prefix="/api")

# JWT Configuration
SECRET_KEY = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"

# ==================== MODELS ====================

# Import Indian currency utilities
from indian_currency import format_indian_currency, inr
from email_service import send_verification_email, send_password_reset_email

class User(BaseModel):
    user_id: str
    email: str
    name: str
    mobile_number: Optional[str] = None
    security_question: Optional[str] = None
    security_answer: Optional[str] = None
    email_verified: bool = False
    verification_token: Optional[str] = None
    picture: Optional[str] = None
    created_at: datetime
    use_single_user_mode: bool = False

class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    mobile_number: Optional[str] = None
    security_question: Optional[str] = None
    security_answer: Optional[str] = None

class UserLogin(BaseModel):
    email: str
    password: str

# Family Member Model
class FamilyMember(BaseModel):
    family_member_id: str
    user_id: str
    name: str
    role: str  # self, spouse, child, parent
    is_active: bool = True
    created_at: datetime

class FamilyMemberCreate(BaseModel):
    name: str
    role: str

# Account Model
class Account(BaseModel):
    account_id: str
    user_id: str
    family_member_id: Optional[str] = None
    name: str
    account_type: str  # bank, cash, upi, credit_card, wallet, investment_account
    ownership_type: str = "individual"  # individual, joint, business
    institution: Optional[str] = None  # bank name, wallet provider
    balance: float = 0.0
    account_number: Optional[str] = None
    color: Optional[str] = None
    icon: Optional[str] = None
    created_at: datetime
    updated_at: datetime

class AccountCreate(BaseModel):
    name: str
    account_type: str
    ownership_type: str = "individual"
    institution: Optional[str] = None
    initial_balance: float = 0.0
    account_number: Optional[str] = None
    family_member_id: Optional[str] = None
    color: Optional[str] = None
    icon: Optional[str] = None

class AccountUpdate(BaseModel):
    name: Optional[str] = None
    account_number: Optional[str] = None
    ownership_type: Optional[str] = None
    institution: Optional[str] = None
    color: Optional[str] = None
    icon: Optional[str] = None

# Income Model
class Income(BaseModel):
    income_id: str
    user_id: str
    family_member_id: Optional[str] = None
    account_id: str
    amount: float
    category: str  # salary, rental, business, other
    sub_category: Optional[str] = None
    source: str
    date: datetime
    notes: Optional[str] = None
    created_at: datetime

class IncomeCreate(BaseModel):
    account_id: str
    amount: float
    category: str
    sub_category: Optional[str] = None
    source: str
    date: str
    notes: Optional[str] = None
    family_member_id: Optional[str] = None

class IncomeUpdate(BaseModel):
    account_id: Optional[str] = None
    amount: Optional[float] = None
    category: Optional[str] = None
    sub_category: Optional[str] = None
    source: Optional[str] = None
    date: Optional[str] = None
    notes: Optional[str] = None

# Expense Model
class Expense(BaseModel):
    expense_id: str
    user_id: str
    family_member_id: Optional[str] = None
    account_id: str
    amount: float
    category: str
    sub_category: Optional[str] = None
    payment_type: str  # cash, bank, credit_card, upi
    description: str
    date: datetime
    notes: Optional[str] = None
    created_at: datetime

class ExpenseCreate(BaseModel):
    account_id: str
    amount: float
    category: str
    sub_category: Optional[str] = None
    payment_type: str
    description: str
    date: str
    notes: Optional[str] = None
    family_member_id: Optional[str] = None

class ExpenseUpdate(BaseModel):
    account_id: Optional[str] = None
    amount: Optional[float] = None
    category: Optional[str] = None
    sub_category: Optional[str] = None
    payment_type: Optional[str] = None
    description: Optional[str] = None
    date: Optional[str] = None
    notes: Optional[str] = None

# ==================== PHASE 2 MODELS ====================

# Credit Card Model
class CreditCardCreate(BaseModel):
    name: str
    card_number_last4: str = ""
    credit_limit: float
    current_outstanding: float = 0.0
    billing_date: int = 1
    due_date: int = 15
    due_time: str = "10:00"  # HH:MM
    interest_rate: float = 0.0
    family_member_id: Optional[str] = None
    # EMI fields
    emis: Optional[List[Dict]] = []  # [{name, amount, tenure_remaining, total_tenure, start_date}]

class CreditCardUpdate(BaseModel):
    name: Optional[str] = None
    credit_limit: Optional[float] = None
    current_outstanding: Optional[float] = None
    billing_date: Optional[int] = None
    due_date: Optional[int] = None
    due_time: Optional[str] = None
    interest_rate: Optional[float] = None
    emis: Optional[List[Dict]] = None

# Loan Model
class LoanCreate(BaseModel):
    name: str
    loan_type: str  # home, car, personal, education, gold, other
    principal_amount: float
    outstanding_amount: float
    interest_rate: float
    emi_amount: float
    tenure_months: int
    start_date: str
    next_emi_date: Optional[str] = None
    account_id: Optional[str] = None
    family_member_id: Optional[str] = None
    notes: Optional[str] = None

class LoanUpdate(BaseModel):
    name: Optional[str] = None
    outstanding_amount: Optional[float] = None
    emi_amount: Optional[float] = None
    next_emi_date: Optional[str] = None
    notes: Optional[str] = None

# Lending (Money Lent / Borrowed) Model
class LendingCreate(BaseModel):
    lending_type: str  # lent, borrowed
    person_name: str
    amount: float
    date: str
    due_date: Optional[str] = None
    notes: Optional[str] = None

class LendingUpdate(BaseModel):
    remaining_amount: Optional[float] = None
    due_date: Optional[str] = None
    notes: Optional[str] = None
    is_settled: Optional[bool] = None

# Investment Model
class InvestmentCreate(BaseModel):
    name: str
    investment_type: str  # stocks, mutual_fund, fd, rd, ppf, nps, gold, real_estate, crypto, other
    invested_amount: float
    current_value: float
    purchase_date: str
    maturity_date: Optional[str] = None
    family_member_id: Optional[str] = None
    notes: Optional[str] = None
    heading_id: Optional[str] = None
    sub_category: Optional[str] = None

class InvestmentUpdate(BaseModel):
    name: Optional[str] = None
    current_value: Optional[float] = None
    maturity_date: Optional[str] = None
    notes: Optional[str] = None

# Reminder Model
class ReminderCreate(BaseModel):
    title: str
    description: Optional[str] = None
    reminder_date: str  # ISO date
    reminder_type: str  # investment, loan_emi, credit_card, lending, bill, custom
    related_id: Optional[str] = None  # Links to investment_id, loan_id, card_id, lending_id, bill_id
    is_recurring: bool = False
    recurrence: Optional[str] = None  # daily, weekly, monthly, yearly

class ReminderUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    reminder_date: Optional[str] = None
    is_completed: Optional[bool] = None
    is_recurring: Optional[bool] = None
    recurrence: Optional[str] = None

# Rental Income Models
class RentalCreate(BaseModel):
    property_name: str
    tenant_name: str = ""
    rent_amount: float
    due_day: int = 1  # Day of month rent is due
    address: str = ""
    notes: str = ""

class RentalUpdate(BaseModel):
    property_name: Optional[str] = None
    tenant_name: Optional[str] = None
    rent_amount: Optional[float] = None
    due_day: Optional[int] = None
    address: Optional[str] = None
    notes: Optional[str] = None

class RentalPaymentCreate(BaseModel):
    rental_id: str
    amount: float
    payment_date: str
    notes: str = ""

# Investment Heading Models
class InvestmentHeadingCreate(BaseModel):
    name: str  # e.g., "Shares", "Mutual Funds"
    icon: str = "trending-up"

class InvestmentHeadingUpdate(BaseModel):
    name: Optional[str] = None
    icon: Optional[str] = None

# Notes Model
class NoteCreate(BaseModel):
    title: str
    content: str = ""
    sections: Optional[list] = None  # [{heading: str, content: str}]
    tags: Optional[list] = None  # ["investment", "tax"]
    linked_type: Optional[str] = None  # transaction, investment, bill
    linked_id: Optional[str] = None
    priority: str = "normal"  # low, normal, high
    color: Optional[str] = None

class NoteUpdate(BaseModel):
    title: Optional[str] = None
    content: Optional[str] = None
    sections: Optional[list] = None
    tags: Optional[list] = None
    linked_type: Optional[str] = None
    linked_id: Optional[str] = None
    priority: Optional[str] = None
    color: Optional[str] = None
    is_archived: Optional[bool] = None

class Bill(BaseModel):
    bill_id: str
    user_id: str
    family_member_id: Optional[str] = None
    account_id: Optional[str] = None
    name: str
    amount: float
    currency: str = "INR"
    due_date: datetime
    category: str
    vendor: Optional[str] = None
    notes: Optional[str] = None
    receipt_image: Optional[str] = None  # base64 encoded
    is_recurring: bool = False
    recurrence_type: Optional[str] = None  # daily, weekly, monthly, yearly
    recurrence_interval: Optional[int] = 1
    payment_status: str = "unpaid"  # unpaid, paid
    created_at: datetime
    updated_at: datetime

class BillCreate(BaseModel):
    name: str
    amount: float
    currency: str = "INR"
    due_date: str
    category: str
    vendor: Optional[str] = None
    notes: Optional[str] = None
    receipt_image: Optional[str] = None
    is_recurring: bool = False
    recurrence_type: Optional[str] = None
    recurrence_interval: Optional[int] = 1
    account_id: Optional[str] = None
    family_member_id: Optional[str] = None

class BillUpdate(BaseModel):
    name: Optional[str] = None
    amount: Optional[float] = None
    currency: Optional[str] = None
    due_date: Optional[str] = None
    category: Optional[str] = None
    vendor: Optional[str] = None
    notes: Optional[str] = None
    receipt_image: Optional[str] = None
    is_recurring: Optional[bool] = None
    recurrence_type: Optional[str] = None
    recurrence_interval: Optional[int] = None
    payment_status: Optional[str] = None
    account_id: Optional[str] = None

class Payment(BaseModel):
    payment_id: str
    bill_id: str
    user_id: str
    amount: float
    payment_date: datetime
    payment_method: Optional[str] = None
    confirmation_number: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime

class PaymentCreate(BaseModel):
    bill_id: str
    amount: float
    payment_date: str
    payment_method: Optional[str] = None
    confirmation_number: Optional[str] = None
    notes: Optional[str] = None

class Budget(BaseModel):
    budget_id: str
    user_id: str
    category: str
    monthly_limit: float
    created_at: datetime
    updated_at: datetime

class BudgetCreate(BaseModel):
    category: str
    monthly_limit: float

class BudgetUpdate(BaseModel):
    monthly_limit: Optional[float] = None
    category: Optional[str] = None

class Category(BaseModel):
    category_id: str
    user_id: str
    name: str
    color: str
    icon: Optional[str] = None
    created_at: datetime

class CategoryCreate(BaseModel):
    name: str
    color: str
    icon: Optional[str] = None

class UserSettings(BaseModel):
    user_id: str
    dark_mode: bool = False
    notifications_enabled: bool = True
    notification_days_before: int = 3
    default_currency: str = "INR"
    storage_provider: str = "local"  # local, google_drive, onedrive
    updated_at: datetime

class SettingsUpdate(BaseModel):
    dark_mode: Optional[bool] = None
    notifications_enabled: Optional[bool] = None
    notification_days_before: Optional[int] = None
    default_currency: Optional[str] = None
    storage_provider: Optional[str] = None

# ==================== HELPER FUNCTIONS ====================

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=7)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def verify_token(token: str):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        logger.info(f"Token verified successfully for user: {payload.get('user_id', 'unknown')}")
        return payload
    except jwt.ExpiredSignatureError as e:
        logger.error(f"Token expired: {e}")
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError as e:
        logger.error(f"Invalid token error: {e}")
        raise HTTPException(status_code=401, detail="Invalid token")
    except Exception as e:
        logger.error(f"Unexpected error verifying token: {type(e).__name__}: {e}")
        raise HTTPException(status_code=401, detail=f"Token verification failed: {str(e)}")

async def get_current_user(request: Request):
    """Get current user from session_token cookie or Authorization header"""
    token = None
    
    # First try to get from cookie
    token = request.cookies.get("session_token")
    
    # Fallback to Authorization header
    if not token:
        auth_header = request.headers.get("authorization") or request.headers.get("Authorization")
        if auth_header:
            if auth_header.startswith("Bearer "):
                token = auth_header[7:]  # Remove "Bearer " prefix
            else:
                token = auth_header
    
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Check if it's a session token (Google OAuth)
    session_doc = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if session_doc:
        # Verify session not expired
        expires_at = session_doc["expires_at"]
        if isinstance(expires_at, str):
            expires_at = datetime.fromisoformat(expires_at)
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
        if expires_at < datetime.now(timezone.utc):
            raise HTTPException(status_code=401, detail="Session expired")
        
        # Get user
        user_doc = await db.users.find_one({"user_id": session_doc["user_id"]}, {"_id": 0, "password_hash": 0})
        if not user_doc:
            raise HTTPException(status_code=404, detail="User not found")
        return User(**user_doc)
    
    # Otherwise verify JWT token
    payload = verify_token(token)
    user_id = payload.get("user_id")
    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    return User(**user_doc)

# ==================== AUTH ENDPOINTS ====================

@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    """Register new user with email/password"""
    # Check if user exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Hash password
    hashed_password = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt())
    
    # Create user
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    verification_token = uuid.uuid4().hex
    user = {
        "user_id": user_id,
        "email": user_data.email,
        "name": user_data.name,
        "mobile_number": user_data.mobile_number,
        "security_question": user_data.security_question,
        "security_answer": user_data.security_answer,
        "password_hash": hashed_password.decode('utf-8'),
        "email_verified": False,
        "verification_token": verification_token,
        "picture": None,
        "created_at": datetime.now(timezone.utc),
        "use_single_user_mode": False
    }
    
    await db.users.insert_one(user)
    
    # Send verification email (non-blocking, don't fail registration if email fails)
    try:
        send_verification_email(user_data.email, user_data.name, verification_token)
    except Exception as e:
        logger.warning(f"Failed to send verification email to {user_data.email}: {e}")
    
    # Create default settings
    settings = {
        "user_id": user_id,
        "dark_mode": False,
        "notifications_enabled": True,
        "notification_days_before": 3,
        "default_currency": "INR",
        "storage_provider": "local",
        "updated_at": datetime.now(timezone.utc)
    }
    await db.user_settings.insert_one(settings)
    
    # Create token
    access_token = create_access_token({"user_id": user_id, "email": user_data.email})
    
    # Return user data without password
    user.pop("password_hash", None)
    user.pop("_id", None)
    
    return {
        "user": user,
        "access_token": access_token,
        "token_type": "bearer"
    }

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    """Login with email/password"""
    user_doc = await db.users.find_one({"email": credentials.email})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Google OAuth users don't have a password — reject email/password login
    if "password_hash" not in user_doc or not user_doc["password_hash"]:
        raise HTTPException(status_code=401, detail="This account uses Google Sign-In. Please login with Google.")
    
    # Verify password
    if not bcrypt.checkpw(credentials.password.encode('utf-8'), user_doc["password_hash"].encode('utf-8')):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Create token
    access_token = create_access_token({"user_id": user_doc["user_id"], "email": user_doc["email"]})
    
    # Return user data without password
    user_doc.pop("password_hash", None)
    user_doc.pop("_id", None)
    
    return {
        "user": user_doc,
        "access_token": access_token,
        "token_type": "bearer"
    }

@api_router.post("/auth/google/session")
async def google_auth_session(request: Request, response: Response):
    """REMINDER: DO NOT HARDCODE THE URL, OR ADD ANY FALLBACKS OR REDIRECT URLS, THIS BREAKS THE AUTH"""
    data = await request.json()
    session_id = data.get("session_id")
    
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    
    # Call Emergent Auth to get session data
    async with httpx.AsyncClient() as client:
        try:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_id},
                timeout=10.0
            )
            if auth_response.status_code != 200:
                raise HTTPException(status_code=401, detail="Invalid session")
            
            session_data = auth_response.json()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Auth service error: {str(e)}")
    
    # Get or create user
    user_doc = await db.users.find_one({"email": session_data["email"]}, {"_id": 0})
    
    if not user_doc:
        # Create new user
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user_doc = {
            "user_id": user_id,
            "email": session_data["email"],
            "name": session_data.get("name", ""),
            "picture": session_data.get("picture"),
            "created_at": datetime.now(timezone.utc),
            "use_single_user_mode": False
        }
        await db.users.insert_one(user_doc)
        
        # Create default settings
        settings = {
            "user_id": user_id,
            "dark_mode": False,
            "notifications_enabled": True,
            "notification_days_before": 3,
            "default_currency": "INR",
            "storage_provider": "local",
            "updated_at": datetime.now(timezone.utc)
        }
        await db.user_settings.insert_one(settings)
        
        user_doc.pop("_id", None)
    else:
        # Update user info if needed
        await db.users.update_one(
            {"user_id": user_doc["user_id"]},
            {"$set": {
                "name": session_data.get("name", user_doc.get("name", "")),
                "picture": session_data.get("picture", user_doc.get("picture"))
            }}
        )
    
    # Store session
    session_token = session_data["session_token"]
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    await db.user_sessions.insert_one({
        "user_id": user_doc["user_id"],
        "session_token": session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc)
    })
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7*24*60*60
    )
    
    return {"user": user_doc, "session_token": session_token}

@api_router.get("/auth/me")
async def get_me(request: Request):
    """Get current authenticated user"""
    user = await get_current_user(request=request)
    return user

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    """Logout user"""
    token = request.cookies.get("session_token")
    if token:
        await db.user_sessions.delete_one({"session_token": token})
        response.delete_cookie("session_token", path="/")
    return {"message": "Logged out successfully"}

# --- Email Verification ---

@api_router.get("/auth/verify-email")
async def verify_email(token: str):
    """Verify email address using the token sent via email"""
    user_doc = await db.users.find_one({"verification_token": token})
    if not user_doc:
        raise HTTPException(status_code=400, detail="Invalid or expired verification token")

    if user_doc.get("email_verified"):
        return {"message": "Email already verified"}

    await db.users.update_one(
        {"user_id": user_doc["user_id"]},
        {"$set": {"email_verified": True, "verification_token": None}}
    )
    return {"message": "Email verified successfully"}

@api_router.post("/auth/resend-verification")
async def resend_verification(request: Request):
    """Resend verification email to the current user"""
    user = await get_current_user(request)
    if user.email_verified:
        return {"message": "Email already verified"}

    new_token = uuid.uuid4().hex
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$set": {"verification_token": new_token}}
    )
    try:
        send_verification_email(user.email, user.name, new_token)
    except Exception as e:
        logger.warning(f"Failed to resend verification email: {e}")
    return {"message": "Verification email sent"}

# --- Password Reset ---

class ForgotPasswordRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

@api_router.post("/auth/forgot-password")
async def forgot_password(data: ForgotPasswordRequest):
    """Send password reset email"""
    user_doc = await db.users.find_one({"email": data.email})
    # Always return success to prevent email enumeration
    if not user_doc:
        return {"message": "If an account exists with that email, a reset link has been sent."}

    # Don't allow reset for Google-only accounts
    if "password_hash" not in user_doc or not user_doc.get("password_hash"):
        return {"message": "If an account exists with that email, a reset link has been sent."}

    reset_token = uuid.uuid4().hex
    expires_at = datetime.now(timezone.utc) + timedelta(hours=1)
    await db.password_resets.insert_one({
        "user_id": user_doc["user_id"],
        "email": data.email,
        "reset_token": reset_token,
        "expires_at": expires_at,
        "used": False,
        "created_at": datetime.now(timezone.utc)
    })

    try:
        send_password_reset_email(data.email, user_doc.get("name", ""), reset_token)
    except Exception as e:
        logger.warning(f"Failed to send password reset email: {e}")

    return {"message": "If an account exists with that email, a reset link has been sent."}

@api_router.post("/auth/reset-password")
async def reset_password(data: ResetPasswordRequest):
    """Reset password using token from email"""
    reset_doc = await db.password_resets.find_one({
        "reset_token": data.token,
        "used": False
    })
    if not reset_doc:
        raise HTTPException(status_code=400, detail="Invalid or expired reset token")

    expires_at = reset_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Reset token has expired")

    # Hash new password and update user
    hashed_password = bcrypt.hashpw(data.new_password.encode('utf-8'), bcrypt.gensalt())
    await db.users.update_one(
        {"user_id": reset_doc["user_id"]},
        {"$set": {"password_hash": hashed_password.decode('utf-8')}}
    )

    # Mark token as used
    await db.password_resets.update_one(
        {"reset_token": data.token},
        {"$set": {"used": True}}
    )

    return {"message": "Password reset successfully"}

@api_router.post("/auth/single-user")
async def single_user_mode():
    """Create or get single-user mode user"""
    # Check if single user exists
    user_doc = await db.users.find_one({"use_single_user_mode": True}, {"_id": 0})
    
    if not user_doc:
        # Create single user
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user_doc = {
            "user_id": user_id,
            "email": "single-user@local",
            "name": "Local User",
            "picture": None,
            "created_at": datetime.now(timezone.utc),
            "use_single_user_mode": True
        }
        await db.users.insert_one(user_doc)
        
        # Create default settings
        settings = {
            "user_id": user_id,
            "dark_mode": False,
            "notifications_enabled": True,
            "notification_days_before": 3,
            "default_currency": "INR",
            "storage_provider": "local",
            "updated_at": datetime.now(timezone.utc)
        }
        await db.user_settings.insert_one(settings)
        
        user_doc.pop("_id", None)
    
    # Create token
    access_token = create_access_token({"user_id": user_doc["user_id"], "email": user_doc["email"]})
    
    return {
        "user": user_doc,
        "access_token": access_token,
        "token_type": "bearer"
    }

# ==================== BILL ENDPOINTS ====================

@api_router.post("/bills", response_model=Bill)
async def create_bill(bill_data: BillCreate, request: Request, authorization: Optional[str] = Header(None)):
    """Create new bill"""
    user = await get_current_user(request)
    
    bill_id = f"bill_{uuid.uuid4().hex[:12]}"
    bill = {
        "bill_id": bill_id,
        "user_id": user.user_id,
        "family_member_id": bill_data.family_member_id,
        "account_id": bill_data.account_id,
        "name": bill_data.name,
        "amount": bill_data.amount,
        "currency": bill_data.currency,
        "due_date": datetime.fromisoformat(bill_data.due_date.replace('Z', '+00:00')),
        "category": bill_data.category,
        "vendor": bill_data.vendor,
        "notes": bill_data.notes,
        "receipt_image": bill_data.receipt_image,
        "is_recurring": bill_data.is_recurring,
        "recurrence_type": bill_data.recurrence_type,
        "recurrence_interval": bill_data.recurrence_interval,
        "payment_status": "unpaid",
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.bills.insert_one(bill)
    bill.pop("_id", None)
    
    return Bill(**bill)

@api_router.get("/bills", response_model=List[Bill])
async def get_bills(
    request: Request,
    authorization: Optional[str] = Header(None),
    month: Optional[int] = None,
    year: Optional[int] = None,
    category: Optional[str] = None,
    status: Optional[str] = None
):
    """Get all bills for user with optional filtering"""
    user = await get_current_user(request)
    
    query = {"user_id": user.user_id}
    
    # Filter by month/year
    if month and year:
        start_date = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        query["due_date"] = {"$gte": start_date, "$lt": end_date}
    
    # Filter by category
    if category:
        query["category"] = category
    
    # Filter by status
    if status:
        query["payment_status"] = status
    
    bills = await db.bills.find(query, {"_id": 0}).sort("due_date", 1).to_list(1000)
    return [Bill(**bill) for bill in bills]

@api_router.get("/bills/summary")
async def bills_summary_early(request: Request):
    """Get bills with overdue/upcoming/paid status"""
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    all_bills = await db.bills.find({"user_id": user.user_id}, {"_id": 0}).sort("due_date", 1).to_list(1000)

    overdue = []
    upcoming = []
    paid_bills = []
    for b in all_bills:
        due = b.get("due_date")
        if isinstance(due, str):
            try:
                due = datetime.fromisoformat(due.replace('Z', '+00:00'))
            except (ValueError, TypeError):
                due = None
        elif isinstance(due, datetime):
            if due.tzinfo is None:
                due = due.replace(tzinfo=timezone.utc)
        status = b.get("payment_status", b.get("status", "pending"))
        if status == "paid":
            paid_bills.append({**b, "bill_status": "paid"})
        elif due and due < now:
            overdue.append({**b, "bill_status": "overdue", "days_overdue": (now - due).days})
        else:
            days_until = (due - now).days if due else 999
            upcoming.append({**b, "bill_status": "upcoming", "days_until": days_until})

    return {
        "overdue": overdue, "overdue_count": len(overdue),
        "upcoming": upcoming, "upcoming_count": len(upcoming),
        "paid": paid_bills, "paid_count": len(paid_bills),
        "total_overdue_amount": sum(b.get("amount", 0) for b in overdue),
        "total_upcoming_amount": sum(b.get("amount", 0) for b in upcoming),
    }

@api_router.get("/bills/{bill_id}", response_model=Bill)
async def get_bill(bill_id: str, request: Request, authorization: Optional[str] = Header(None)):
    """Get specific bill"""
    user = await get_current_user(request)
    
    bill_doc = await db.bills.find_one({"bill_id": bill_id, "user_id": user.user_id}, {"_id": 0})
    if not bill_doc:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    return Bill(**bill_doc)

@api_router.put("/bills/{bill_id}", response_model=Bill)
async def update_bill(bill_id: str, bill_data: BillUpdate, request: Request, authorization: Optional[str] = Header(None)):
    """Update bill"""
    user = await get_current_user(request)
    
    # Check bill exists and belongs to user
    existing_bill = await db.bills.find_one({"bill_id": bill_id, "user_id": user.user_id})
    if not existing_bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    # Prepare update data
    update_data = {k: v for k, v in bill_data.model_dump(exclude_unset=True).items() if v is not None}
    if "due_date" in update_data:
        update_data["due_date"] = datetime.fromisoformat(update_data["due_date"].replace('Z', '+00:00'))
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # Update bill
    await db.bills.update_one(
        {"bill_id": bill_id, "user_id": user.user_id},
        {"$set": update_data}
    )
    
    # Get updated bill
    updated_bill = await db.bills.find_one({"bill_id": bill_id}, {"_id": 0})
    return Bill(**updated_bill)

@api_router.delete("/bills/{bill_id}")
async def delete_bill(bill_id: str, request: Request, authorization: Optional[str] = Header(None)):
    """Delete bill"""
    user = await get_current_user(request)
    
    result = await db.bills.delete_one({"bill_id": bill_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    # Also delete associated payments
    await db.payments.delete_many({"bill_id": bill_id})
    
    return {"message": "Bill deleted successfully"}

# ==================== PAYMENT ENDPOINTS ====================

@api_router.post("/payments", response_model=Payment)
async def create_payment(payment_data: PaymentCreate, request: Request, authorization: Optional[str] = Header(None)):
    """Record a payment"""
    user = await get_current_user(request)
    
    # Verify bill exists and belongs to user
    bill = await db.bills.find_one({"bill_id": payment_data.bill_id, "user_id": user.user_id})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    
    payment_id = f"payment_{uuid.uuid4().hex[:12]}"
    payment = {
        "payment_id": payment_id,
        "bill_id": payment_data.bill_id,
        "user_id": user.user_id,
        "amount": payment_data.amount,
        "payment_date": datetime.fromisoformat(payment_data.payment_date.replace('Z', '+00:00')),
        "payment_method": payment_data.payment_method,
        "confirmation_number": payment_data.confirmation_number,
        "notes": payment_data.notes,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.payments.insert_one(payment)
    
    # Update bill status to paid
    await db.bills.update_one(
        {"bill_id": payment_data.bill_id},
        {"$set": {"payment_status": "paid", "updated_at": datetime.now(timezone.utc)}}
    )
    
    payment.pop("_id", None)
    return Payment(**payment)

@api_router.get("/payments", response_model=List[Payment])
async def get_payments(request: Request, authorization: Optional[str] = Header(None), bill_id: Optional[str] = None):
    """Get payment history"""
    user = await get_current_user(request)
    
    query = {"user_id": user.user_id}
    if bill_id:
        query["bill_id"] = bill_id
    
    payments = await db.payments.find(query, {"_id": 0}).sort("payment_date", -1).to_list(1000)
    return [Payment(**payment) for payment in payments]

# ==================== CATEGORY ENDPOINTS ====================

@api_router.post("/categories", response_model=Category)
async def create_category(category_data: CategoryCreate, request: Request, authorization: Optional[str] = Header(None)):
    """Create custom category"""
    user = await get_current_user(request)
    
    # Check if category already exists
    existing = await db.categories.find_one({"user_id": user.user_id, "name": category_data.name})
    if existing:
        raise HTTPException(status_code=400, detail="Category already exists")
    
    category_id = f"cat_{uuid.uuid4().hex[:12]}"
    category = {
        "category_id": category_id,
        "user_id": user.user_id,
        "name": category_data.name,
        "color": category_data.color,
        "icon": category_data.icon,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.categories.insert_one(category)
    category.pop("_id", None)
    
    return Category(**category)

@api_router.get("/categories", response_model=List[Category])
async def get_categories(request: Request, authorization: Optional[str] = Header(None)):
    """Get all categories"""
    user = await get_current_user(request)
    
    categories = await db.categories.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    return [Category(**cat) for cat in categories]

# ==================== BUDGET ENDPOINTS ====================

@api_router.post("/budgets", response_model=Budget)
async def create_budget(budget_data: BudgetCreate, request: Request, authorization: Optional[str] = Header(None)):
    """Set budget limit for category"""
    user = await get_current_user(request)
    
    # Check if budget already exists for this category
    existing = await db.budgets.find_one({"user_id": user.user_id, "category": budget_data.category})
    if existing:
        # Update existing budget
        await db.budgets.update_one(
            {"user_id": user.user_id, "category": budget_data.category},
            {"$set": {"monthly_limit": budget_data.monthly_limit, "updated_at": datetime.now(timezone.utc)}}
        )
        updated_budget = await db.budgets.find_one({"user_id": user.user_id, "category": budget_data.category}, {"_id": 0})
        return Budget(**updated_budget)
    
    budget_id = f"budget_{uuid.uuid4().hex[:12]}"
    budget = {
        "budget_id": budget_id,
        "user_id": user.user_id,
        "category": budget_data.category,
        "monthly_limit": budget_data.monthly_limit,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.budgets.insert_one(budget)
    budget.pop("_id", None)
    
    return Budget(**budget)

@api_router.get("/budgets", response_model=List[Budget])
async def get_budgets(request: Request, authorization: Optional[str] = Header(None)):
    """Get all budgets"""
    user = await get_current_user(request)
    
    budgets = await db.budgets.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    return [Budget(**budget) for budget in budgets]

@api_router.get("/budgets/progress")
async def get_budget_progress(request: Request, month: Optional[int] = None, year: Optional[int] = None):
    """Get budget progress with actual spending per category"""
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    m = month or now.month
    y = year or now.year

    budgets = await db.budgets.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)

    # Get expenses for this month
    expenses = await db.expenses.find({
        "user_id": user.user_id,
        "date": {"$gte": datetime(y, m, 1, tzinfo=timezone.utc),
                 "$lt": datetime(y + (1 if m == 12 else 0), (m % 12) + 1, 1, tzinfo=timezone.utc)}
    }, {"_id": 0}).to_list(50000)

    # Calculate spending per category
    spending: Dict = {}
    for exp in expenses:
        cat = exp.get("category", "other")
        spending[cat] = spending.get(cat, 0) + exp.get("amount", 0)

    total_budgeted = 0
    total_spent = 0
    progress = []
    for b in budgets:
        cat = b["category"]
        limit = b["monthly_limit"]
        spent = spending.get(cat, 0)
        total_budgeted += limit
        total_spent += spent
        pct = round((spent / limit * 100), 1) if limit > 0 else 0
        remaining = limit - spent
        status = "over_budget" if spent > limit else "warning" if pct > 80 else "on_track"
        progress.append({
            "budget_id": b["budget_id"],
            "category": cat,
            "monthly_limit": limit,
            "spent": spent,
            "remaining": remaining,
            "percentage": pct,
            "status": status,
        })

    progress.sort(key=lambda x: x["percentage"], reverse=True)

    # Unbudgeted spending (categories with expenses but no budget)
    unbudgeted = []
    budgeted_cats = {b["category"] for b in budgets}
    for cat, amount in spending.items():
        if cat not in budgeted_cats:
            unbudgeted.append({"category": cat, "spent": amount})

    return {
        "month": m, "year": y,
        "total_budgeted": total_budgeted,
        "total_spent": total_spent,
        "overall_percentage": round((total_spent / total_budgeted * 100), 1) if total_budgeted > 0 else 0,
        "budgets": progress,
        "unbudgeted_spending": sorted(unbudgeted, key=lambda x: x["spent"], reverse=True),
    }

@api_router.put("/budgets/{budget_id}")
async def update_budget(budget_id: str, data: BudgetUpdate, request: Request):
    """Update a budget"""
    user = await get_current_user(request)
    existing = await db.budgets.find_one({"budget_id": budget_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Budget not found")
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    await db.budgets.update_one({"budget_id": budget_id}, {"$set": update_data})
    updated = await db.budgets.find_one({"budget_id": budget_id}, {"_id": 0})
    return Budget(**updated)

@api_router.delete("/budgets/{budget_id}")
async def delete_budget(budget_id: str, request: Request):
    """Delete a budget"""
    user = await get_current_user(request)
    result = await db.budgets.delete_one({"budget_id": budget_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Budget not found")
    return {"message": "Budget deleted"}

# ==================== ANALYTICS ENDPOINTS ====================

@api_router.get("/analytics/spending")
async def get_spending_analytics(
    request: Request,
    authorization: Optional[str] = Header(None),
    month: Optional[int] = None,
    year: Optional[int] = None
):
    """Get spending analytics"""
    user = await get_current_user(request)
    
    # Default to current month
    if not month or not year:
        now = datetime.now(timezone.utc)
        month = now.month
        year = now.year
    
    # Get all bills for the month
    start_date = datetime(year, month, 1, tzinfo=timezone.utc)
    if month == 12:
        end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
    
    bills = await db.bills.find({
        "user_id": user.user_id,
        "due_date": {"$gte": start_date, "$lt": end_date}
    }, {"_id": 0}).to_list(1000)
    
    # Calculate analytics
    total_amount = sum(bill["amount"] for bill in bills)
    paid_amount = sum(bill["amount"] for bill in bills if bill["payment_status"] == "paid")
    unpaid_amount = total_amount - paid_amount
    
    # Category breakdown
    category_totals = {}
    for bill in bills:
        category = bill["category"]
        if category not in category_totals:
            category_totals[category] = 0
        category_totals[category] += bill["amount"]
    
    # Budget comparison
    budgets = await db.budgets.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    budget_status = []
    for budget in budgets:
        spent = category_totals.get(budget["category"], 0)
        budget_status.append({
            "category": budget["category"],
            "limit": budget["monthly_limit"],
            "spent": spent,
            "remaining": budget["monthly_limit"] - spent,
            "percentage": (spent / budget["monthly_limit"] * 100) if budget["monthly_limit"] > 0 else 0
        })
    
    return {
        "month": month,
        "year": year,
        "total_amount": total_amount,
        "paid_amount": paid_amount,
        "unpaid_amount": unpaid_amount,
        "total_bills": len(bills),
        "paid_bills": len([b for b in bills if b["payment_status"] == "paid"]),
        "unpaid_bills": len([b for b in bills if b["payment_status"] == "unpaid"]),
        "category_breakdown": [{"category": k, "amount": v} for k, v in category_totals.items()],
        "budget_status": budget_status
    }

# ==================== SETTINGS ENDPOINTS ====================

@api_router.get("/settings", response_model=UserSettings)
async def get_settings(request: Request, authorization: Optional[str] = Header(None)):
    """Get user settings"""
    user = await get_current_user(request)
    
    settings = await db.user_settings.find_one({"user_id": user.user_id}, {"_id": 0})
    if not settings:
        # Create default settings
        settings = {
            "user_id": user.user_id,
            "dark_mode": False,
            "notifications_enabled": True,
            "notification_days_before": 3,
            "default_currency": "INR",
            "storage_provider": "local",
            "updated_at": datetime.now(timezone.utc)
        }
        await db.user_settings.insert_one(settings)
        settings.pop("_id", None)
    
    return UserSettings(**settings)

@api_router.put("/settings", response_model=UserSettings)
async def update_settings(settings_data: SettingsUpdate, request: Request, authorization: Optional[str] = Header(None)):
    """Update user settings"""
    user = await get_current_user(request)
    
    update_data = {k: v for k, v in settings_data.model_dump(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.user_settings.update_one(
        {"user_id": user.user_id},
        {"$set": update_data},
        upsert=True
    )
    
    settings = await db.user_settings.find_one({"user_id": user.user_id}, {"_id": 0})
    return UserSettings(**settings)

# ==================== FAMILY MEMBER ENDPOINTS ====================

@api_router.post("/family-members")
async def create_family_member(data: FamilyMemberCreate, request: Request):
    """Create a new family member"""
    user = await get_current_user(request)
    
    member_id = f"fm_{uuid.uuid4().hex[:12]}"
    member = {
        "family_member_id": member_id,
        "user_id": user.user_id,
        "name": data.name,
        "role": data.role,
        "is_active": True,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.family_members.insert_one(member)
    member.pop("_id", None)
    return member

@api_router.get("/family-members")
async def get_family_members(request: Request):
    """Get all family members for user"""
    user = await get_current_user(request)
    
    members = await db.family_members.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    return members

@api_router.put("/family-members/{member_id}")
async def update_family_member(member_id: str, data: FamilyMemberCreate, request: Request):
    """Update a family member"""
    user = await get_current_user(request)
    
    existing = await db.family_members.find_one(
        {"family_member_id": member_id, "user_id": user.user_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Family member not found")
    
    update_data = {"name": data.name, "role": data.role}
    await db.family_members.update_one(
        {"family_member_id": member_id, "user_id": user.user_id},
        {"$set": update_data}
    )
    
    updated = await db.family_members.find_one(
        {"family_member_id": member_id}, {"_id": 0}
    )
    return updated

@api_router.delete("/family-members/{member_id}")
async def delete_family_member(member_id: str, request: Request):
    """Delete a family member"""
    user = await get_current_user(request)
    
    result = await db.family_members.delete_one(
        {"family_member_id": member_id, "user_id": user.user_id}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Family member not found")
    
    return {"message": "Family member deleted successfully"}

# ==================== ACCOUNT ENDPOINTS ====================

@api_router.post("/accounts")
async def create_account(data: AccountCreate, request: Request):
    """Create a new financial account"""
    user = await get_current_user(request)
    
    # Validate family member if provided
    if data.family_member_id:
        fm = await db.family_members.find_one(
            {"family_member_id": data.family_member_id, "user_id": user.user_id}
        )
        if not fm:
            raise HTTPException(status_code=404, detail="Family member not found")
    
    account_id = f"acc_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    account = {
        "account_id": account_id,
        "user_id": user.user_id,
        "family_member_id": data.family_member_id,
        "name": data.name,
        "account_type": data.account_type,
        "ownership_type": data.ownership_type or "individual",
        "institution": data.institution,
        "balance": data.initial_balance,
        "account_number": data.account_number,
        "color": data.color,
        "icon": data.icon,
        "is_active": True,
        "created_at": now,
        "updated_at": now
    }
    
    await db.accounts.insert_one(account)
    account.pop("_id", None)
    return account

@api_router.get("/accounts")
async def get_accounts(
    request: Request,
    account_type: Optional[str] = None,
    family_member_id: Optional[str] = None
):
    """Get all accounts for user with optional filters"""
    user = await get_current_user(request)
    
    query = {"user_id": user.user_id, "is_active": True}
    if account_type:
        query["account_type"] = account_type
    if family_member_id:
        query["family_member_id"] = family_member_id
    
    accounts = await db.accounts.find(query, {"_id": 0}).sort("created_at", 1).to_list(100)
    return accounts

@api_router.get("/accounts/{account_id}")
async def get_account(account_id: str, request: Request):
    """Get a specific account"""
    user = await get_current_user(request)
    
    account = await db.accounts.find_one(
        {"account_id": account_id, "user_id": user.user_id}, {"_id": 0}
    )
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    return account

@api_router.put("/accounts/{account_id}")
async def update_account(account_id: str, data: AccountUpdate, request: Request):
    """Update an account"""
    user = await get_current_user(request)
    
    existing = await db.accounts.find_one(
        {"account_id": account_id, "user_id": user.user_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Account not found")
    
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    await db.accounts.update_one(
        {"account_id": account_id, "user_id": user.user_id},
        {"$set": update_data}
    )
    
    updated = await db.accounts.find_one({"account_id": account_id}, {"_id": 0})
    return updated

@api_router.delete("/accounts/{account_id}")
async def delete_account(account_id: str, request: Request):
    """Soft-delete an account (preserves transaction history)"""
    user = await get_current_user(request)
    
    existing = await db.accounts.find_one(
        {"account_id": account_id, "user_id": user.user_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Soft delete — mark inactive instead of removing
    await db.accounts.update_one(
        {"account_id": account_id, "user_id": user.user_id},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": "Account deactivated successfully"}

# ==================== INCOME ENDPOINTS ====================

@api_router.post("/income")
async def create_income(data: IncomeCreate, request: Request):
    """Create an income entry and update account balance"""
    user = await get_current_user(request)
    
    # Validate account
    account = await db.accounts.find_one(
        {"account_id": data.account_id, "user_id": user.user_id, "is_active": True}
    )
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Validate family member if provided
    if data.family_member_id:
        fm = await db.family_members.find_one(
            {"family_member_id": data.family_member_id, "user_id": user.user_id}
        )
        if not fm:
            raise HTTPException(status_code=404, detail="Family member not found")
    
    income_id = f"inc_{uuid.uuid4().hex[:12]}"
    income = {
        "income_id": income_id,
        "user_id": user.user_id,
        "family_member_id": data.family_member_id,
        "account_id": data.account_id,
        "amount": data.amount,
        "category": data.category,
        "sub_category": data.sub_category,
        "source": data.source,
        "date": datetime.fromisoformat(data.date.replace('Z', '+00:00')),
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.income.insert_one(income)
    
    # Update account balance (+)
    await db.accounts.update_one(
        {"account_id": data.account_id},
        {"$inc": {"balance": data.amount}, "$set": {"updated_at": datetime.now(timezone.utc)}}
    )
    
    income.pop("_id", None)
    return income

@api_router.get("/income")
async def get_income(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    account_id: Optional[str] = None,
    family_member_id: Optional[str] = None,
    month: Optional[int] = None,
    year: Optional[int] = None
):
    """Get income entries with advanced filters"""
    user = await get_current_user(request)
    
    query = {"user_id": user.user_id}
    
    # Date range filter
    if start_date and end_date:
        query["date"] = {
            "$gte": datetime.fromisoformat(start_date.replace('Z', '+00:00')),
            "$lte": datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        }
    elif month and year:
        start = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        query["date"] = {"$gte": start, "$lt": end}
    
    if category:
        query["category"] = category
    if account_id:
        query["account_id"] = account_id
    if family_member_id:
        query["family_member_id"] = family_member_id
    
    incomes = await db.income.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return incomes

@api_router.get("/income/{income_id}")
async def get_income_single(income_id: str, request: Request):
    """Get a specific income entry"""
    user = await get_current_user(request)
    
    income = await db.income.find_one(
        {"income_id": income_id, "user_id": user.user_id}, {"_id": 0}
    )
    if not income:
        raise HTTPException(status_code=404, detail="Income entry not found")
    return income

@api_router.put("/income/{income_id}")
async def update_income(income_id: str, data: IncomeUpdate, request: Request):
    """Update an income entry and adjust account balance"""
    user = await get_current_user(request)
    
    existing = await db.income.find_one(
        {"income_id": income_id, "user_id": user.user_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Income entry not found")
    
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    
    # Handle amount change → adjust account balance
    if "amount" in update_data:
        old_amount = existing["amount"]
        new_amount = update_data["amount"]
        diff = new_amount - old_amount
        
        account_id = update_data.get("account_id", existing["account_id"])
        
        # If account changed, reverse from old, add to new
        if "account_id" in update_data and update_data["account_id"] != existing["account_id"]:
            await db.accounts.update_one(
                {"account_id": existing["account_id"]},
                {"$inc": {"balance": -old_amount}, "$set": {"updated_at": datetime.now(timezone.utc)}}
            )
            await db.accounts.update_one(
                {"account_id": update_data["account_id"]},
                {"$inc": {"balance": new_amount}, "$set": {"updated_at": datetime.now(timezone.utc)}}
            )
        else:
            await db.accounts.update_one(
                {"account_id": existing["account_id"]},
                {"$inc": {"balance": diff}, "$set": {"updated_at": datetime.now(timezone.utc)}}
            )
    elif "account_id" in update_data and update_data["account_id"] != existing["account_id"]:
        # Account changed but amount didn't
        await db.accounts.update_one(
            {"account_id": existing["account_id"]},
            {"$inc": {"balance": -existing["amount"]}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
        await db.accounts.update_one(
            {"account_id": update_data["account_id"]},
            {"$inc": {"balance": existing["amount"]}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
    
    if "date" in update_data:
        update_data["date"] = datetime.fromisoformat(update_data["date"].replace('Z', '+00:00'))
    
    await db.income.update_one(
        {"income_id": income_id, "user_id": user.user_id},
        {"$set": update_data}
    )
    
    updated = await db.income.find_one({"income_id": income_id}, {"_id": 0})
    return updated

@api_router.delete("/income/{income_id}")
async def delete_income(income_id: str, request: Request):
    """Delete an income entry and reverse account balance"""
    user = await get_current_user(request)
    
    existing = await db.income.find_one(
        {"income_id": income_id, "user_id": user.user_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Income entry not found")
    
    # Reverse the balance change
    await db.accounts.update_one(
        {"account_id": existing["account_id"]},
        {"$inc": {"balance": -existing["amount"]}, "$set": {"updated_at": datetime.now(timezone.utc)}}
    )
    
    await db.income.delete_one({"income_id": income_id, "user_id": user.user_id})
    return {"message": "Income entry deleted successfully"}

# ==================== EXPENSE ENDPOINTS ====================

@api_router.post("/expenses")
async def create_expense(data: ExpenseCreate, request: Request):
    """Create an expense entry and update account balance"""
    user = await get_current_user(request)
    
    # Validate account
    account = await db.accounts.find_one(
        {"account_id": data.account_id, "user_id": user.user_id, "is_active": True}
    )
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Validate family member if provided
    if data.family_member_id:
        fm = await db.family_members.find_one(
            {"family_member_id": data.family_member_id, "user_id": user.user_id}
        )
        if not fm:
            raise HTTPException(status_code=404, detail="Family member not found")
    
    expense_id = f"exp_{uuid.uuid4().hex[:12]}"
    expense = {
        "expense_id": expense_id,
        "user_id": user.user_id,
        "family_member_id": data.family_member_id,
        "account_id": data.account_id,
        "amount": data.amount,
        "category": data.category,
        "sub_category": data.sub_category,
        "payment_type": data.payment_type,
        "description": data.description,
        "date": datetime.fromisoformat(data.date.replace('Z', '+00:00')),
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.expenses.insert_one(expense)
    
    # Update account balance (-)
    await db.accounts.update_one(
        {"account_id": data.account_id},
        {"$inc": {"balance": -data.amount}, "$set": {"updated_at": datetime.now(timezone.utc)}}
    )
    
    expense.pop("_id", None)
    return expense

@api_router.get("/expenses")
async def get_expenses(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    category: Optional[str] = None,
    account_id: Optional[str] = None,
    family_member_id: Optional[str] = None,
    payment_type: Optional[str] = None,
    month: Optional[int] = None,
    year: Optional[int] = None
):
    """Get expense entries with advanced filters"""
    user = await get_current_user(request)
    
    query = {"user_id": user.user_id}
    
    # Date range filter
    if start_date and end_date:
        query["date"] = {
            "$gte": datetime.fromisoformat(start_date.replace('Z', '+00:00')),
            "$lte": datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        }
    elif month and year:
        start = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end = datetime(year, month + 1, 1, tzinfo=timezone.utc)
        query["date"] = {"$gte": start, "$lt": end}
    
    if category:
        query["category"] = category
    if account_id:
        query["account_id"] = account_id
    if family_member_id:
        query["family_member_id"] = family_member_id
    if payment_type:
        query["payment_type"] = payment_type
    
    expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return expenses

@api_router.get("/expenses/{expense_id}")
async def get_expense_single(expense_id: str, request: Request):
    """Get a specific expense entry"""
    user = await get_current_user(request)
    
    expense = await db.expenses.find_one(
        {"expense_id": expense_id, "user_id": user.user_id}, {"_id": 0}
    )
    if not expense:
        raise HTTPException(status_code=404, detail="Expense entry not found")
    return expense

@api_router.put("/expenses/{expense_id}")
async def update_expense(expense_id: str, data: ExpenseUpdate, request: Request):
    """Update an expense entry and adjust account balance"""
    user = await get_current_user(request)
    
    existing = await db.expenses.find_one(
        {"expense_id": expense_id, "user_id": user.user_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Expense entry not found")
    
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    
    # Handle amount change → adjust account balance
    if "amount" in update_data:
        old_amount = existing["amount"]
        new_amount = update_data["amount"]
        diff = new_amount - old_amount
        
        if "account_id" in update_data and update_data["account_id"] != existing["account_id"]:
            # Account changed: reverse old, apply new
            await db.accounts.update_one(
                {"account_id": existing["account_id"]},
                {"$inc": {"balance": old_amount}, "$set": {"updated_at": datetime.now(timezone.utc)}}
            )
            await db.accounts.update_one(
                {"account_id": update_data["account_id"]},
                {"$inc": {"balance": -new_amount}, "$set": {"updated_at": datetime.now(timezone.utc)}}
            )
        else:
            # Same account: adjust by diff (negative because expense)
            await db.accounts.update_one(
                {"account_id": existing["account_id"]},
                {"$inc": {"balance": -diff}, "$set": {"updated_at": datetime.now(timezone.utc)}}
            )
    elif "account_id" in update_data and update_data["account_id"] != existing["account_id"]:
        # Account changed but amount didn't
        await db.accounts.update_one(
            {"account_id": existing["account_id"]},
            {"$inc": {"balance": existing["amount"]}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
        await db.accounts.update_one(
            {"account_id": update_data["account_id"]},
            {"$inc": {"balance": -existing["amount"]}, "$set": {"updated_at": datetime.now(timezone.utc)}}
        )
    
    if "date" in update_data:
        update_data["date"] = datetime.fromisoformat(update_data["date"].replace('Z', '+00:00'))
    
    await db.expenses.update_one(
        {"expense_id": expense_id, "user_id": user.user_id},
        {"$set": update_data}
    )
    
    updated = await db.expenses.find_one({"expense_id": expense_id}, {"_id": 0})
    return updated

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, request: Request):
    """Delete an expense entry and reverse account balance"""
    user = await get_current_user(request)
    
    existing = await db.expenses.find_one(
        {"expense_id": expense_id, "user_id": user.user_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Expense entry not found")
    
    # Reverse the balance change (add back the expense amount)
    await db.accounts.update_one(
        {"account_id": existing["account_id"]},
        {"$inc": {"balance": existing["amount"]}, "$set": {"updated_at": datetime.now(timezone.utc)}}
    )
    
    await db.expenses.delete_one({"expense_id": expense_id, "user_id": user.user_id})
    return {"message": "Expense entry deleted successfully"}

# ==================== CREDIT CARD ENDPOINTS ====================

@api_router.post("/credit-cards")
async def create_credit_card(data: CreditCardCreate, request: Request):
    user = await get_current_user(request)
    card_id = f"cc_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    card = {
        "card_id": card_id, "user_id": user.user_id, "name": data.name,
        "card_number_last4": data.card_number_last4, "credit_limit": data.credit_limit,
        "current_outstanding": data.current_outstanding, "billing_date": data.billing_date,
        "due_date": data.due_date, "due_time": data.due_time, "interest_rate": data.interest_rate,
        "family_member_id": data.family_member_id, "emis": data.emis or [],
        "is_active": True, "created_at": now, "updated_at": now
    }
    await db.credit_cards.insert_one(card)
    card.pop("_id", None)
    return card

@api_router.get("/credit-cards")
async def get_credit_cards(request: Request):
    user = await get_current_user(request)
    cards = await db.credit_cards.find(
        {"user_id": user.user_id, "is_active": True}, {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    return cards

@api_router.get("/credit-cards/report")
async def credit_card_report(request: Request, period: str = "monthly"):
    """Credit card reporting - day/month/year wise"""
    user = await get_current_user(request)
    cards = await db.credit_cards.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(100)
    now = datetime.now(timezone.utc)

    total_limit = sum(c.get("credit_limit", 0) for c in cards)
    total_outstanding = sum(c.get("current_outstanding", 0) for c in cards)
    total_available = total_limit - total_outstanding
    total_emi = sum(sum(e.get("amount", 0) for e in c.get("emis", [])) for c in cards)

    # Upcoming due dates
    upcoming_dues = []
    for c in cards:
        due_day = c.get("due_date", 15)
        due_time = c.get("due_time", "10:00")
        # Next due date — clamp due_day to the last day of the target month
        _, max_day_this_month = calendar.monthrange(now.year, now.month)
        clamped_day = min(due_day, max_day_this_month)
        if now.day <= clamped_day:
            next_due = now.replace(day=clamped_day)
        else:
            next_month = now.month + 1 if now.month < 12 else 1
            next_year = now.year if now.month < 12 else now.year + 1
            _, max_day_next_month = calendar.monthrange(next_year, next_month)
            next_due = now.replace(year=next_year, month=next_month, day=min(due_day, max_day_next_month))
        days_until = (next_due - now).days
        status = "overdue" if days_until < 0 else "critical" if days_until <= 3 else "warning" if days_until <= 7 else "safe"
        upcoming_dues.append({
            "card_id": c["card_id"], "name": c["name"],
            "outstanding": c.get("current_outstanding", 0),
            "due_day": due_day, "due_time": due_time,
            "next_due_date": str(next_due.date()),
            "days_until": days_until, "status": status,
        })
    upcoming_dues.sort(key=lambda x: x["days_until"])

    return {
        "summary": {
            "total_cards": len(cards),
            "total_limit": total_limit,
            "total_outstanding": total_outstanding,
            "total_available": total_available,
            "total_emi": total_emi,
            "utilization": round(total_outstanding / total_limit * 100, 1) if total_limit > 0 else 0,
        },
        "upcoming_dues": upcoming_dues,
        "cards": cards,
    }

@api_router.put("/credit-cards/{card_id}")
async def update_credit_card(card_id: str, data: CreditCardUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.credit_cards.find_one({"card_id": card_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Credit card not found")
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    await db.credit_cards.update_one({"card_id": card_id}, {"$set": update_data})
    updated = await db.credit_cards.find_one({"card_id": card_id}, {"_id": 0})
    return updated

@api_router.delete("/credit-cards/{card_id}")
async def delete_credit_card(card_id: str, request: Request):
    user = await get_current_user(request)
    existing = await db.credit_cards.find_one({"card_id": card_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Credit card not found")
    await db.credit_cards.update_one({"card_id": card_id}, {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}})
    return {"message": "Credit card deactivated"}

# ==================== LOAN ENDPOINTS ====================

@api_router.post("/loans")
async def create_loan(data: LoanCreate, request: Request):
    user = await get_current_user(request)
    loan_id = f"loan_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    loan = {
        "loan_id": loan_id, "user_id": user.user_id, "name": data.name,
        "loan_type": data.loan_type, "principal_amount": data.principal_amount,
        "outstanding_amount": data.outstanding_amount, "interest_rate": data.interest_rate,
        "emi_amount": data.emi_amount, "tenure_months": data.tenure_months,
        "start_date": datetime.fromisoformat(data.start_date.replace('Z', '+00:00')),
        "next_emi_date": datetime.fromisoformat(data.next_emi_date.replace('Z', '+00:00')) if data.next_emi_date else None,
        "account_id": data.account_id, "family_member_id": data.family_member_id,
        "notes": data.notes, "is_active": True, "created_at": now, "updated_at": now
    }
    await db.loans.insert_one(loan)
    loan.pop("_id", None)
    return loan

@api_router.get("/loans")
async def get_loans(request: Request):
    user = await get_current_user(request)
    loans = await db.loans.find(
        {"user_id": user.user_id, "is_active": True}, {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    return loans

@api_router.put("/loans/{loan_id}")
async def update_loan(loan_id: str, data: LoanUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.loans.find_one({"loan_id": loan_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Loan not found")
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if "next_emi_date" in update_data and isinstance(update_data["next_emi_date"], str):
        update_data["next_emi_date"] = datetime.fromisoformat(update_data["next_emi_date"].replace('Z', '+00:00'))
    update_data["updated_at"] = datetime.now(timezone.utc)
    await db.loans.update_one({"loan_id": loan_id}, {"$set": update_data})
    updated = await db.loans.find_one({"loan_id": loan_id}, {"_id": 0})
    return updated

@api_router.delete("/loans/{loan_id}")
async def delete_loan(loan_id: str, request: Request):
    user = await get_current_user(request)
    existing = await db.loans.find_one({"loan_id": loan_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Loan not found")
    await db.loans.update_one({"loan_id": loan_id}, {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc)}})
    return {"message": "Loan deactivated"}

# ==================== LENDING (LENT/BORROWED) ENDPOINTS ====================

@api_router.post("/lending")
async def create_lending(data: LendingCreate, request: Request):
    user = await get_current_user(request)
    lending_id = f"lend_{uuid.uuid4().hex[:12]}"
    lending = {
        "lending_id": lending_id, "user_id": user.user_id,
        "lending_type": data.lending_type, "person_name": data.person_name,
        "amount": data.amount, "remaining_amount": data.amount,
        "date": datetime.fromisoformat(data.date.replace('Z', '+00:00')),
        "due_date": datetime.fromisoformat(data.due_date.replace('Z', '+00:00')) if data.due_date else None,
        "notes": data.notes, "is_settled": False,
        "created_at": datetime.now(timezone.utc)
    }
    await db.lending.insert_one(lending)
    lending.pop("_id", None)
    return lending

@api_router.get("/lending")
async def get_lending(request: Request, lending_type: Optional[str] = None, is_settled: Optional[bool] = None):
    user = await get_current_user(request)
    query: Dict = {"user_id": user.user_id}
    if lending_type:
        query["lending_type"] = lending_type
    if is_settled is not None:
        query["is_settled"] = is_settled
    records = await db.lending.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return records

@api_router.put("/lending/{lending_id}")
async def update_lending(lending_id: str, data: LendingUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.lending.find_one({"lending_id": lending_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Lending record not found")
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if "due_date" in update_data and isinstance(update_data["due_date"], str):
        update_data["due_date"] = datetime.fromisoformat(update_data["due_date"].replace('Z', '+00:00'))
    await db.lending.update_one({"lending_id": lending_id}, {"$set": update_data})
    updated = await db.lending.find_one({"lending_id": lending_id}, {"_id": 0})
    return updated

@api_router.delete("/lending/{lending_id}")
async def delete_lending(lending_id: str, request: Request):
    user = await get_current_user(request)
    result = await db.lending.delete_one({"lending_id": lending_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lending record not found")
    return {"message": "Lending record deleted"}

# ==================== INVESTMENT ENDPOINTS ====================

@api_router.post("/investments")
async def create_investment(data: InvestmentCreate, request: Request):
    user = await get_current_user(request)
    inv_id = f"inv_{uuid.uuid4().hex[:12]}"
    investment = {
        "investment_id": inv_id, "user_id": user.user_id, "name": data.name,
        "investment_type": data.investment_type, "invested_amount": data.invested_amount,
        "current_value": data.current_value,
        "purchase_date": datetime.fromisoformat(data.purchase_date.replace('Z', '+00:00')),
        "maturity_date": datetime.fromisoformat(data.maturity_date.replace('Z', '+00:00')) if data.maturity_date else None,
        "family_member_id": data.family_member_id, "notes": data.notes,
        "heading_id": data.heading_id, "sub_category": data.sub_category,
        "is_active": True, "created_at": datetime.now(timezone.utc)
    }
    await db.investments.insert_one(investment)
    investment.pop("_id", None)
    return investment

@api_router.get("/investments")
async def get_investments(request: Request, investment_type: Optional[str] = None):
    user = await get_current_user(request)
    query: Dict = {"user_id": user.user_id, "is_active": True}
    if investment_type:
        query["investment_type"] = investment_type
    investments = await db.investments.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return investments

@api_router.put("/investments/{inv_id}")
async def update_investment(inv_id: str, data: InvestmentUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.investments.find_one({"investment_id": inv_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Investment not found")
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if "maturity_date" in update_data and isinstance(update_data["maturity_date"], str):
        update_data["maturity_date"] = datetime.fromisoformat(update_data["maturity_date"].replace('Z', '+00:00'))
    await db.investments.update_one({"investment_id": inv_id}, {"$set": update_data})
    updated = await db.investments.find_one({"investment_id": inv_id}, {"_id": 0})
    return updated

@api_router.delete("/investments/{inv_id}")
async def delete_investment(inv_id: str, request: Request):
    user = await get_current_user(request)
    existing = await db.investments.find_one({"investment_id": inv_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Investment not found")
    await db.investments.update_one({"investment_id": inv_id}, {"$set": {"is_active": False}})
    return {"message": "Investment removed"}

# ==================== REMINDERS ENDPOINTS ====================

@api_router.post("/reminders")
async def create_reminder(data: ReminderCreate, request: Request):
    user = await get_current_user(request)
    reminder_id = f"rem_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    reminder = {
        "reminder_id": reminder_id, "user_id": user.user_id,
        "title": data.title, "description": data.description,
        "reminder_date": datetime.fromisoformat(data.reminder_date.replace('Z', '+00:00')),
        "reminder_type": data.reminder_type,
        "related_id": data.related_id,
        "is_recurring": data.is_recurring,
        "recurrence": data.recurrence,
        "is_completed": False,
        "created_at": now, "updated_at": now
    }
    await db.reminders.insert_one(reminder)
    reminder.pop("_id", None)
    return reminder

@api_router.get("/reminders")
async def get_reminders(request: Request, reminder_type: Optional[str] = None, is_completed: Optional[bool] = None, upcoming: Optional[bool] = None):
    user = await get_current_user(request)
    query: Dict = {"user_id": user.user_id}
    if reminder_type:
        query["reminder_type"] = reminder_type
    if is_completed is not None:
        query["is_completed"] = is_completed
    if upcoming:
        query["reminder_date"] = {"$gte": datetime.now(timezone.utc)}
        query["is_completed"] = False
    reminders = await db.reminders.find(query, {"_id": 0}).sort("reminder_date", 1).to_list(1000)
    # Enrich with related item details
    for r in reminders:
        if r.get("related_id"):
            if r["reminder_type"] == "investment":
                item = await db.investments.find_one({"investment_id": r["related_id"]}, {"_id": 0, "name": 1, "investment_type": 1, "current_value": 1})
                r["related_item"] = item
            elif r["reminder_type"] == "loan_emi":
                item = await db.loans.find_one({"loan_id": r["related_id"]}, {"_id": 0, "name": 1, "loan_type": 1, "emi_amount": 1})
                r["related_item"] = item
            elif r["reminder_type"] == "credit_card":
                item = await db.credit_cards.find_one({"card_id": r["related_id"]}, {"_id": 0, "name": 1, "current_outstanding": 1})
                r["related_item"] = item
            elif r["reminder_type"] == "lending":
                item = await db.lending.find_one({"lending_id": r["related_id"]}, {"_id": 0, "person_name": 1, "lending_type": 1, "remaining_amount": 1})
                r["related_item"] = item
            elif r["reminder_type"] == "bill":
                item = await db.bills.find_one({"bill_id": r["related_id"]}, {"_id": 0, "name": 1, "amount": 1})
                r["related_item"] = item
    return reminders

@api_router.get("/reminders/summary")
async def get_reminders_summary(request: Request):
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    week_end = today_start + timedelta(days=7)
    
    total = await db.reminders.count_documents({"user_id": user.user_id, "is_completed": False})
    overdue = await db.reminders.count_documents({"user_id": user.user_id, "is_completed": False, "reminder_date": {"$lt": now}})
    today = await db.reminders.count_documents({"user_id": user.user_id, "is_completed": False, "reminder_date": {"$gte": today_start, "$lt": today_end}})
    this_week = await db.reminders.count_documents({"user_id": user.user_id, "is_completed": False, "reminder_date": {"$gte": today_start, "$lt": week_end}})
    
    upcoming_list = await db.reminders.find(
        {"user_id": user.user_id, "is_completed": False, "reminder_date": {"$gte": now}},
        {"_id": 0}
    ).sort("reminder_date", 1).to_list(5)
    
    overdue_list = await db.reminders.find(
        {"user_id": user.user_id, "is_completed": False, "reminder_date": {"$lt": now}},
        {"_id": 0}
    ).sort("reminder_date", -1).to_list(5)
    
    return {
        "total_pending": total,
        "overdue": overdue,
        "today": today,
        "this_week": this_week,
        "upcoming": upcoming_list,
        "overdue_list": overdue_list
    }

@api_router.put("/reminders/{reminder_id}")
async def update_reminder(reminder_id: str, data: ReminderUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.reminders.find_one({"reminder_id": reminder_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Reminder not found")
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    if "reminder_date" in update_data and isinstance(update_data["reminder_date"], str):
        update_data["reminder_date"] = datetime.fromisoformat(update_data["reminder_date"].replace('Z', '+00:00'))
    update_data["updated_at"] = datetime.now(timezone.utc)
    await db.reminders.update_one({"reminder_id": reminder_id}, {"$set": update_data})
    updated = await db.reminders.find_one({"reminder_id": reminder_id}, {"_id": 0})
    return updated

@api_router.delete("/reminders/{reminder_id}")
async def delete_reminder(reminder_id: str, request: Request):
    user = await get_current_user(request)
    result = await db.reminders.delete_one({"reminder_id": reminder_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Reminder not found")
    return {"message": "Reminder deleted"}

# ==================== RENTAL INCOME ENDPOINTS ====================

@api_router.post("/rentals")
async def create_rental(data: RentalCreate, request: Request):
    user = await get_current_user(request)
    rental_id = f"rent_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    rental = {
        "rental_id": rental_id, "user_id": user.user_id,
        "property_name": data.property_name, "tenant_name": data.tenant_name,
        "rent_amount": data.rent_amount, "due_day": data.due_day,
        "address": data.address, "notes": data.notes,
        "is_active": True, "payments": [],
        "created_at": now, "updated_at": now
    }
    await db.rentals.insert_one(rental)
    rental.pop("_id", None)
    return rental

@api_router.get("/rentals")
async def get_rentals(request: Request):
    user = await get_current_user(request)
    rentals = await db.rentals.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).sort("created_at", -1).to_list(100)
    # Enrich with payment status for current month
    now = datetime.now(timezone.utc)
    for r in rentals:
        payments = r.get("payments", [])
        current_month_paid = any(
            p.get("month") == now.month and p.get("year") == now.year for p in payments
        )
        r["current_month_paid"] = current_month_paid
        r["total_collected"] = sum(p.get("amount", 0) for p in payments)
    return rentals

@api_router.post("/rentals/{rental_id}/payments")
async def add_rental_payment(rental_id: str, data: RentalPaymentCreate, request: Request):
    user = await get_current_user(request)
    rental = await db.rentals.find_one({"rental_id": rental_id, "user_id": user.user_id})
    if not rental:
        raise HTTPException(status_code=404, detail="Rental not found")
    payment_date = datetime.fromisoformat(data.payment_date.replace('Z', '+00:00'))
    payment = {
        "payment_id": f"rp_{uuid.uuid4().hex[:8]}",
        "amount": data.amount, "payment_date": payment_date,
        "month": payment_date.month, "year": payment_date.year,
        "notes": data.notes
    }
    await db.rentals.update_one({"rental_id": rental_id}, {"$push": {"payments": payment}, "$set": {"updated_at": datetime.now(timezone.utc)}})
    return payment

@api_router.put("/rentals/{rental_id}")
async def update_rental(rental_id: str, data: RentalUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.rentals.find_one({"rental_id": rental_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Rental not found")
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    await db.rentals.update_one({"rental_id": rental_id}, {"$set": update_data})
    updated = await db.rentals.find_one({"rental_id": rental_id}, {"_id": 0})
    return updated

@api_router.delete("/rentals/{rental_id}")
async def delete_rental(rental_id: str, request: Request):
    user = await get_current_user(request)
    result = await db.rentals.delete_one({"rental_id": rental_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rental not found")
    return {"message": "Rental deleted"}

# ==================== INVESTMENT HEADINGS ENDPOINTS ====================

@api_router.post("/investment-headings")
async def create_heading(data: InvestmentHeadingCreate, request: Request):
    user = await get_current_user(request)
    heading_id = f"ih_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    heading = {
        "heading_id": heading_id, "user_id": user.user_id,
        "name": data.name, "icon": data.icon,
        "created_at": now, "updated_at": now
    }
    await db.investment_headings.insert_one(heading)
    heading.pop("_id", None)
    return heading

@api_router.get("/investment-headings")
async def get_headings(request: Request):
    user = await get_current_user(request)
    headings = await db.investment_headings.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    # Enrich with investments under each heading
    for h in headings:
        investments = await db.investments.find({"user_id": user.user_id, "heading_id": h["heading_id"], "is_active": True}, {"_id": 0}).to_list(1000)
        h["investments"] = investments
        h["total_invested"] = sum(i.get("invested_amount", 0) for i in investments)
        h["total_current"] = sum(i.get("current_value", 0) for i in investments)
        h["count"] = len(investments)
    return headings

@api_router.put("/investment-headings/{heading_id}")
async def update_heading(heading_id: str, data: InvestmentHeadingUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.investment_headings.find_one({"heading_id": heading_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Heading not found")
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    await db.investment_headings.update_one({"heading_id": heading_id}, {"$set": update_data})
    updated = await db.investment_headings.find_one({"heading_id": heading_id}, {"_id": 0})
    return updated

@api_router.delete("/investment-headings/{heading_id}")
async def delete_heading(heading_id: str, request: Request):
    user = await get_current_user(request)
    result = await db.investment_headings.delete_one({"heading_id": heading_id, "user_id": user.user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Heading not found")
    return {"message": "Heading deleted"}

# ==================== NET WORTH ENDPOINT ====================

@api_router.get("/net-worth")
async def get_net_worth(request: Request):
    user = await get_current_user(request)
    
    # Assets
    accounts = await db.accounts.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(100)
    total_account_balance = sum(a.get("balance", 0) for a in accounts)
    
    investments = await db.investments.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(1000)
    total_investment_value = sum(i.get("current_value", 0) for i in investments)
    total_invested = sum(i.get("invested_amount", 0) for i in investments)
    
    lent_records = await db.lending.find({"user_id": user.user_id, "lending_type": "lent", "is_settled": False}, {"_id": 0}).to_list(1000)
    total_lent = sum(l.get("remaining_amount", 0) for l in lent_records)
    
    total_assets = total_account_balance + total_investment_value + total_lent
    
    # Liabilities
    credit_cards = await db.credit_cards.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(100)
    total_cc_outstanding = sum(c.get("current_outstanding", 0) for c in credit_cards)
    
    loans = await db.loans.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(100)
    total_loan_outstanding = sum(l.get("outstanding_amount", 0) for l in loans)
    
    borrowed_records = await db.lending.find({"user_id": user.user_id, "lending_type": "borrowed", "is_settled": False}, {"_id": 0}).to_list(1000)
    total_borrowed = sum(b.get("remaining_amount", 0) for b in borrowed_records)
    
    total_liabilities = total_cc_outstanding + total_loan_outstanding + total_borrowed
    net_worth = total_assets - total_liabilities
    
    return {
        "net_worth": net_worth,
        "net_worth_formatted": format_indian_currency(net_worth),
        "total_assets": total_assets,
        "total_assets_formatted": format_indian_currency(total_assets),
        "total_liabilities": total_liabilities,
        "total_liabilities_formatted": format_indian_currency(total_liabilities),
        "assets": {
            "accounts": {"total": total_account_balance, "formatted": format_indian_currency(total_account_balance), "items": accounts},
            "investments": {"total": total_investment_value, "invested": total_invested, "formatted": format_indian_currency(total_investment_value), "items": investments},
            "money_lent": {"total": total_lent, "formatted": format_indian_currency(total_lent), "items": lent_records},
        },
        "liabilities": {
            "credit_cards": {"total": total_cc_outstanding, "formatted": format_indian_currency(total_cc_outstanding), "items": credit_cards},
            "loans": {"total": total_loan_outstanding, "formatted": format_indian_currency(total_loan_outstanding), "items": loans},
            "money_borrowed": {"total": total_borrowed, "formatted": format_indian_currency(total_borrowed), "items": borrowed_records},
        }
    }

# ==================== DASHBOARD ENDPOINT ====================

@api_router.get("/dashboard")
async def get_dashboard(request: Request):
    """Get financial dashboard summary"""
    user = await get_current_user(request)
    
    now = datetime.now(timezone.utc)
    month_start = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
    if now.month == 12:
        month_end = datetime(now.year + 1, 1, 1, tzinfo=timezone.utc)
    else:
        month_end = datetime(now.year, now.month + 1, 1, tzinfo=timezone.utc)
    
    # Total balance across all active accounts
    accounts = await db.accounts.find(
        {"user_id": user.user_id, "is_active": True}, {"_id": 0}
    ).to_list(100)
    total_balance = sum(a.get("balance", 0) for a in accounts)
    
    # Monthly income
    monthly_incomes = await db.income.find(
        {"user_id": user.user_id, "date": {"$gte": month_start, "$lt": month_end}},
        {"_id": 0}
    ).to_list(1000)
    total_monthly_income = sum(i["amount"] for i in monthly_incomes)
    
    # Monthly expenses
    monthly_expenses = await db.expenses.find(
        {"user_id": user.user_id, "date": {"$gte": month_start, "$lt": month_end}},
        {"_id": 0}
    ).to_list(1000)
    total_monthly_expenses = sum(e["amount"] for e in monthly_expenses)
    
    # Upcoming bills (unpaid, due within 30 days)
    upcoming_bills = await db.bills.find(
        {
            "user_id": user.user_id,
            "payment_status": "unpaid",
            "due_date": {"$gte": now, "$lte": now + timedelta(days=30)}
        },
        {"_id": 0}
    ).sort("due_date", 1).to_list(10)
    
    # Overdue bills
    overdue_bills = await db.bills.find(
        {
            "user_id": user.user_id,
            "payment_status": "unpaid",
            "due_date": {"$lt": now}
        },
        {"_id": 0}
    ).sort("due_date", 1).to_list(10)
    
    # Recent transactions (last 10 combined income + expenses)
    recent_incomes = await db.income.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("date", -1).to_list(10)
    
    recent_expenses = await db.expenses.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("date", -1).to_list(10)
    
    # Merge and sort recent transactions
    recent_transactions = []
    for inc in recent_incomes:
        recent_transactions.append({
            "type": "income",
            "id": inc["income_id"],
            "amount": inc["amount"],
            "category": inc["category"],
            "description": inc.get("source", ""),
            "date": inc["date"],
            "account_id": inc["account_id"]
        })
    for exp in recent_expenses:
        recent_transactions.append({
            "type": "expense",
            "id": exp["expense_id"],
            "amount": exp["amount"],
            "category": exp["category"],
            "description": exp.get("description", ""),
            "date": exp["date"],
            "account_id": exp["account_id"]
        })
    
    recent_transactions.sort(key=lambda x: x["date"], reverse=True)
    recent_transactions = recent_transactions[:10]
    
    # Income category breakdown for the month
    income_by_category = {}
    for inc in monthly_incomes:
        cat = inc["category"]
        income_by_category[cat] = income_by_category.get(cat, 0) + inc["amount"]
    
    # Expense category breakdown for the month
    expense_by_category = {}
    for exp in monthly_expenses:
        cat = exp["category"]
        expense_by_category[cat] = expense_by_category.get(cat, 0) + exp["amount"]
    
    # Account-wise summary
    account_summary = []
    for acc in accounts:
        account_summary.append({
            "account_id": acc["account_id"],
            "name": acc["name"],
            "account_type": acc["account_type"],
            "balance": acc["balance"]
        })
    
    # Family members
    family_members = await db.family_members.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).to_list(50)
    
    return {
        "total_balance": total_balance,
        "total_balance_formatted": format_indian_currency(total_balance),
        "monthly_income": total_monthly_income,
        "monthly_income_formatted": format_indian_currency(total_monthly_income),
        "monthly_expenses": total_monthly_expenses,
        "monthly_expenses_formatted": format_indian_currency(total_monthly_expenses),
        "monthly_savings": total_monthly_income - total_monthly_expenses,
        "monthly_savings_formatted": format_indian_currency(total_monthly_income - total_monthly_expenses),
        "accounts": account_summary,
        "upcoming_bills": upcoming_bills,
        "overdue_bills": overdue_bills,
        "recent_transactions": recent_transactions,
        "income_by_category": [{"category": k, "amount": v} for k, v in income_by_category.items()],
        "expense_by_category": [{"category": k, "amount": v} for k, v in expense_by_category.items()],
        "family_members": family_members,
        "month": now.month,
        "year": now.year
    }

# ==================== ANALYTICS ENDPOINTS ====================

@api_router.get("/analytics/investment")
async def investment_analytics(request: Request):
    """CAGR, portfolio allocation, top/bottom performers"""
    user = await get_current_user(request)
    investments = await db.investments.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(1000)

    total_invested = sum(i.get("invested_amount", 0) for i in investments)
    total_current = sum(i.get("current_value", 0) for i in investments)
    total_returns = total_current - total_invested
    total_returns_pct = (total_returns / total_invested * 100) if total_invested > 0 else 0

    # Portfolio allocation by type
    allocation = {}
    for inv in investments:
        t = inv.get("investment_type", "other")
        if t not in allocation:
            allocation[t] = {"type": t, "invested": 0, "current": 0, "count": 0}
        allocation[t]["invested"] += inv.get("invested_amount", 0)
        allocation[t]["current"] += inv.get("current_value", 0)
        allocation[t]["count"] += 1
    for k in allocation:
        allocation[k]["percentage"] = round(allocation[k]["current"] / total_current * 100, 1) if total_current > 0 else 0
        allocation[k]["returns"] = allocation[k]["current"] - allocation[k]["invested"]
        allocation[k]["returns_pct"] = round(allocation[k]["returns"] / allocation[k]["invested"] * 100, 1) if allocation[k]["invested"] > 0 else 0

    # CAGR per investment
    performers = []
    now = datetime.now(timezone.utc)
    for inv in investments:
        invested = inv.get("invested_amount", 0)
        current = inv.get("current_value", 0)
        ret = current - invested
        ret_pct = (ret / invested * 100) if invested > 0 else 0
        # CAGR calculation
        purchase_date = inv.get("purchase_date")
        cagr = 0
        if purchase_date and invested > 0 and current > 0:
            if isinstance(purchase_date, str):
                try:
                    purchase_date = datetime.fromisoformat(purchase_date.replace('Z', '+00:00'))
                except (ValueError, TypeError):
                    purchase_date = None
            if purchase_date:
                # Ensure both datetimes have timezone info for comparison
                if purchase_date.tzinfo is None:
                    purchase_date = purchase_date.replace(tzinfo=timezone.utc)
                years = max((now - purchase_date).days / 365.25, 0.01)
                try:
                    cagr = round((math.pow(current / invested, 1 / years) - 1) * 100, 2)
                except (ValueError, ZeroDivisionError, OverflowError):
                    cagr = 0
        performers.append({
            "investment_id": inv.get("investment_id"),
            "name": inv.get("name"),
            "type": inv.get("investment_type"),
            "invested": invested,
            "current": current,
            "returns": ret,
            "returns_pct": round(ret_pct, 2),
            "cagr": cagr,
            "purchase_date": str(inv.get("purchase_date", "")),
        })

    performers.sort(key=lambda x: x["returns_pct"], reverse=True)

    return {
        "summary": {
            "total_invested": total_invested,
            "total_current": total_current,
            "total_returns": total_returns,
            "total_returns_pct": round(total_returns_pct, 2),
            "total_investments": len(investments),
        },
        "allocation": list(allocation.values()),
        "top_performers": performers[:5],
        "bottom_performers": list(reversed(performers))[:5] if performers else [],
        "all_performers": performers,
    }


@api_router.get("/analytics/cashflow")
async def cashflow_analytics(request: Request, months: int = 6):
    """Monthly income vs expense trend and savings rate"""
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)

    monthly_data = []
    for i in range(months - 1, -1, -1):
        # Calculate exact month by subtracting i months from current month
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1

        income_list = await db.income.find({
            "user_id": user.user_id,
            "date": {"$gte": datetime(y, m, 1, tzinfo=timezone.utc),
                     "$lt": datetime(y + (1 if m == 12 else 0), (m % 12) + 1, 1, tzinfo=timezone.utc)}
        }, {"_id": 0}).to_list(10000)

        expense_list = await db.expenses.find({
            "user_id": user.user_id,
            "date": {"$gte": datetime(y, m, 1, tzinfo=timezone.utc),
                     "$lt": datetime(y + (1 if m == 12 else 0), (m % 12) + 1, 1, tzinfo=timezone.utc)}
        }, {"_id": 0}).to_list(10000)

        total_income = sum(item.get("amount", 0) for item in income_list)
        total_expense = sum(item.get("amount", 0) for item in expense_list)
        savings = total_income - total_expense
        savings_rate = round((savings / total_income * 100), 1) if total_income > 0 else 0

        month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        monthly_data.append({
            "month": m,
            "year": y,
            "label": f"{month_names[m - 1]} {y}",
            "short_label": month_names[m - 1],
            "income": total_income,
            "expense": total_expense,
            "savings": savings,
            "savings_rate": savings_rate,
        })

    total_income_all = sum(d["income"] for d in monthly_data)
    total_expense_all = sum(d["expense"] for d in monthly_data)
    avg_savings_rate = round((total_income_all - total_expense_all) / total_income_all * 100, 1) if total_income_all > 0 else 0

    return {
        "monthly": monthly_data,
        "summary": {
            "total_income": total_income_all,
            "total_expense": total_expense_all,
            "total_savings": total_income_all - total_expense_all,
            "avg_savings_rate": avg_savings_rate,
        }
    }


@api_router.get("/analytics/expense-breakdown")
async def expense_breakdown(request: Request, month: Optional[int] = None, year: Optional[int] = None):
    """Category-wise expense breakdown"""
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    m = month or now.month
    y = year or now.year

    expenses = await db.expenses.find({
        "user_id": user.user_id,
        "date": {"$gte": datetime(y, m, 1, tzinfo=timezone.utc),
                 "$lt": datetime(y + (1 if m == 12 else 0), (m % 12) + 1, 1, tzinfo=timezone.utc)}
    }, {"_id": 0}).to_list(10000)

    categories: Dict = {}
    total = 0
    for exp in expenses:
        cat = exp.get("category", "other")
        amt = exp.get("amount", 0)
        total += amt
        if cat not in categories:
            categories[cat] = {"category": cat, "amount": 0, "count": 0}
        categories[cat]["amount"] += amt
        categories[cat]["count"] += 1

    for k in categories:
        categories[k]["percentage"] = round(categories[k]["amount"] / total * 100, 1) if total > 0 else 0

    sorted_cats = sorted(categories.values(), key=lambda x: x["amount"], reverse=True)

    return {"total": total, "month": m, "year": y, "categories": sorted_cats}


@api_router.get("/analytics/income-breakdown")
async def income_breakdown(request: Request, month: Optional[int] = None, year: Optional[int] = None):
    """Source-wise income breakdown"""
    user = await get_current_user(request)
    now = datetime.now(timezone.utc)
    m = month or now.month
    y = year or now.year

    incomes = await db.income.find({
        "user_id": user.user_id,
        "date": {"$gte": datetime(y, m, 1, tzinfo=timezone.utc),
                 "$lt": datetime(y + (1 if m == 12 else 0), (m % 12) + 1, 1, tzinfo=timezone.utc)}
    }, {"_id": 0}).to_list(10000)

    categories: Dict = {}
    total = 0
    for inc in incomes:
        cat = inc.get("category", "other")
        amt = inc.get("amount", 0)
        total += amt
        if cat not in categories:
            categories[cat] = {"category": cat, "amount": 0, "count": 0}
        categories[cat]["amount"] += amt
        categories[cat]["count"] += 1

    for k in categories:
        categories[k]["percentage"] = round(categories[k]["amount"] / total * 100, 1) if total > 0 else 0

    sorted_cats = sorted(categories.values(), key=lambda x: x["amount"], reverse=True)

    return {"total": total, "month": m, "year": y, "categories": sorted_cats}


# ==================== CSV EXPORT ENDPOINTS ====================

@api_router.get("/export/transactions-csv")
async def export_transactions_csv(request: Request):
    """Export transactions as CSV"""
    user = await get_current_user(request)
    incomes = await db.income.find({"user_id": user.user_id}, {"_id": 0}).sort("date", -1).to_list(50000)
    expenses = await db.expenses.find({"user_id": user.user_id}, {"_id": 0}).sort("date", -1).to_list(50000)
    accounts = {a["account_id"]: a["name"] for a in await db.accounts.find({"user_id": user.user_id}, {"_id": 0, "account_id": 1, "name": 1}).to_list(100)}

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Type", "Category", "Description", "Amount", "Account", "Payment Type", "Notes"])
    for inc in incomes:
        writer.writerow([str(inc.get("date", "")), "Income", inc.get("category", ""), inc.get("source", ""), inc.get("amount", 0), accounts.get(inc.get("account_id", ""), ""), "", inc.get("notes", "")])
    for exp in expenses:
        writer.writerow([str(exp.get("date", "")), "Expense", exp.get("category", ""), exp.get("description", ""), exp.get("amount", 0), accounts.get(exp.get("account_id", ""), ""), exp.get("payment_type", ""), exp.get("notes", "")])

    return Response(content=output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=transactions.csv"})


@api_router.get("/export/investments-csv")
async def export_investments_csv(request: Request):
    """Export investments as CSV"""
    user = await get_current_user(request)
    investments = await db.investments.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(10000)
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Name", "Type", "Invested Amount", "Current Value", "Returns", "Returns %", "Purchase Date", "Maturity Date", "Notes"])
    for inv in investments:
        invested = inv.get("invested_amount", 0)
        current = inv.get("current_value", 0)
        ret = current - invested
        ret_pct = round(ret / invested * 100, 2) if invested > 0 else 0
        writer.writerow([inv.get("name", ""), inv.get("investment_type", ""), invested, current, ret, f"{ret_pct}%", str(inv.get("purchase_date", "")), str(inv.get("maturity_date", "")), inv.get("notes", "")])
    return Response(content=output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=investments.csv"})


@api_router.get("/export/networth-csv")
async def export_networth_csv(request: Request):
    """Export net worth breakdown as CSV"""
    user = await get_current_user(request)
    accounts = await db.accounts.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(100)
    investments = await db.investments.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(1000)
    credit_cards = await db.credit_cards.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(100)
    loans = await db.loans.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(100)
    lending = await db.lending.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Category", "Item", "Type", "Amount"])
    for a in accounts:
        writer.writerow(["Asset - Account", a.get("name", ""), a.get("account_type", ""), a.get("balance", 0)])
    for i in investments:
        writer.writerow(["Asset - Investment", i.get("name", ""), i.get("investment_type", ""), i.get("current_value", 0)])
    for l in lending:
        if l.get("lending_type") == "lent" and not l.get("is_settled"):
            writer.writerow(["Asset - Money Lent", l.get("person_name", ""), "lent", l.get("remaining_amount", l.get("amount", 0))])
    for c in credit_cards:
        writer.writerow(["Liability - Credit Card", c.get("name", ""), "credit_card", c.get("current_outstanding", 0)])
    for lo in loans:
        writer.writerow(["Liability - Loan", lo.get("name", ""), lo.get("loan_type", ""), lo.get("outstanding_amount", 0)])
    for l in lending:
        if l.get("lending_type") == "borrowed" and not l.get("is_settled"):
            writer.writerow(["Liability - Borrowed", l.get("person_name", ""), "borrowed", l.get("remaining_amount", l.get("amount", 0))])
    return Response(content=output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=networth.csv"})


# ==================== EXPORT ENDPOINT ====================

@api_router.get("/export")
async def export_data(
    request: Request,
    authorization: Optional[str] = Header(None),
    export_format: str = "json"
):
    """Export all user data"""
    user = await get_current_user(request)
    
    # Get all user data
    bills = await db.bills.find({"user_id": user.user_id}, {"_id": 0}).to_list(10000)
    payments = await db.payments.find({"user_id": user.user_id}, {"_id": 0}).to_list(10000)
    categories = await db.categories.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    budgets = await db.budgets.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)
    accounts = await db.accounts.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    incomes = await db.income.find({"user_id": user.user_id}, {"_id": 0}).to_list(10000)
    expenses = await db.expenses.find({"user_id": user.user_id}, {"_id": 0}).to_list(10000)
    family_members = await db.family_members.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    settings = await db.user_settings.find_one({"user_id": user.user_id}, {"_id": 0})
    
    result = {
        "user": user.model_dump(),
        "accounts": accounts,
        "income": incomes,
        "expenses": expenses,
        "bills": bills,
        "payments": payments,
        "categories": categories,
        "budgets": budgets,
        "family_members": family_members,
        "settings": settings,
        "exported_at": datetime.now(timezone.utc).isoformat()
    }
    
    return result


# ==================== EXCEL EXPORT ENDPOINTS ====================

async def _build_transactions_xlsx(user_id: str) -> BytesIO:
    """Build an Excel workbook with transactions data"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    incomes = await db.income.find({"user_id": user_id}, {"_id": 0}).sort("date", -1).to_list(50000)
    expenses = await db.expenses.find({"user_id": user_id}, {"_id": 0}).sort("date", -1).to_list(50000)
    accounts_list = await db.accounts.find({"user_id": user_id}, {"_id": 0, "account_id": 1, "name": 1}).to_list(100)
    accounts_map = {a["account_id"]: a["name"] for a in accounts_list}

    wb = openpyxl.Workbook()

    # --- Styles ---
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    income_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    expense_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC")
    )
    currency_fmt = '#,##0.00'

    # --- All Transactions sheet ---
    ws = wb.active
    ws.title = "All Transactions"
    headers = ["Date", "Type", "Category", "Sub-Category", "Description", "Amount (INR)", "Account", "Payment Type", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    row = 2
    for inc in incomes:
        date_val = inc.get("date", "")
        if isinstance(date_val, datetime):
            date_val = date_val.strftime("%Y-%m-%d")
        ws.cell(row=row, column=1, value=str(date_val)).border = thin_border
        ws.cell(row=row, column=2, value="Income").border = thin_border
        ws.cell(row=row, column=3, value=inc.get("category", "")).border = thin_border
        ws.cell(row=row, column=4, value=inc.get("sub_category", "")).border = thin_border
        ws.cell(row=row, column=5, value=inc.get("source", "")).border = thin_border
        amt_cell = ws.cell(row=row, column=6, value=inc.get("amount", 0))
        amt_cell.number_format = currency_fmt
        amt_cell.border = thin_border
        ws.cell(row=row, column=7, value=accounts_map.get(inc.get("account_id", ""), "")).border = thin_border
        ws.cell(row=row, column=8, value="").border = thin_border
        ws.cell(row=row, column=9, value=inc.get("notes", "") or "").border = thin_border
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = income_fill
        row += 1

    for exp in expenses:
        date_val = exp.get("date", "")
        if isinstance(date_val, datetime):
            date_val = date_val.strftime("%Y-%m-%d")
        ws.cell(row=row, column=1, value=str(date_val)).border = thin_border
        ws.cell(row=row, column=2, value="Expense").border = thin_border
        ws.cell(row=row, column=3, value=exp.get("category", "")).border = thin_border
        ws.cell(row=row, column=4, value=exp.get("sub_category", "")).border = thin_border
        ws.cell(row=row, column=5, value=exp.get("description", "")).border = thin_border
        amt_cell = ws.cell(row=row, column=6, value=exp.get("amount", 0))
        amt_cell.number_format = currency_fmt
        amt_cell.border = thin_border
        ws.cell(row=row, column=7, value=accounts_map.get(exp.get("account_id", ""), "")).border = thin_border
        ws.cell(row=row, column=8, value=exp.get("payment_type", "")).border = thin_border
        ws.cell(row=row, column=9, value=exp.get("notes", "") or "").border = thin_border
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = expense_fill
        row += 1

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 35)

    # --- Income Summary sheet ---
    ws2 = wb.create_sheet("Income Summary")
    ws2.append(["Category", "Total Amount (INR)", "Count"])
    for c in ws2[1]:
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
    income_cats = {}
    for inc in incomes:
        cat = inc.get("category", "other")
        income_cats.setdefault(cat, {"amount": 0, "count": 0})
        income_cats[cat]["amount"] += inc.get("amount", 0)
        income_cats[cat]["count"] += 1
    for cat, data in sorted(income_cats.items(), key=lambda x: x[1]["amount"], reverse=True):
        r = ws2.append([cat, data["amount"], data["count"]])

    # --- Expense Summary sheet ---
    ws3 = wb.create_sheet("Expense Summary")
    ws3.append(["Category", "Total Amount (INR)", "Count"])
    for c in ws3[1]:
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
    expense_cats = {}
    for exp in expenses:
        cat = exp.get("category", "other")
        expense_cats.setdefault(cat, {"amount": 0, "count": 0})
        expense_cats[cat]["amount"] += exp.get("amount", 0)
        expense_cats[cat]["count"] += 1
    for cat, data in sorted(expense_cats.items(), key=lambda x: x[1]["amount"], reverse=True):
        ws3.append([cat, data["amount"], data["count"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@api_router.get("/export/transactions-xlsx")
async def export_transactions_xlsx(request: Request):
    """Export transactions as Excel workbook with formatted sheets"""
    user = await get_current_user(request)
    buf = await _build_transactions_xlsx(user.user_id)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=transactions.xlsx"}
    )


@api_router.get("/export/investments-xlsx")
async def export_investments_xlsx(request: Request):
    """Export investments as Excel workbook"""
    user = await get_current_user(request)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    investments = await db.investments.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Investments"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC")
    )
    green_font = Font(name="Arial", color="27AE60")
    red_font = Font(name="Arial", color="C0392B")
    currency_fmt = '#,##0.00'

    headers = ["Name", "Type", "Invested (INR)", "Current Value (INR)", "Returns (INR)", "Returns %", "CAGR %", "Purchase Date", "Maturity Date", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    now = datetime.now(timezone.utc)
    row = 2
    for inv in investments:
        invested = inv.get("invested_amount", 0)
        current = inv.get("current_value", 0)
        ret = current - invested
        ret_pct = round(ret / invested * 100, 2) if invested > 0 else 0

        # CAGR
        purchase_date = inv.get("purchase_date")
        cagr = 0
        if purchase_date and invested > 0 and current > 0:
            if isinstance(purchase_date, str):
                try:
                    purchase_date = datetime.fromisoformat(purchase_date.replace('Z', '+00:00'))
                except (ValueError, TypeError):
                    purchase_date = None
            if purchase_date:
                if isinstance(purchase_date, datetime) and purchase_date.tzinfo is None:
                    purchase_date = purchase_date.replace(tzinfo=timezone.utc)
                years = max((now - purchase_date).days / 365.25, 0.01)
                try:
                    cagr = round((math.pow(current / invested, 1 / years) - 1) * 100, 2)
                except (ValueError, ZeroDivisionError, OverflowError):
                    cagr = 0

        pd_val = inv.get("purchase_date", "")
        if isinstance(pd_val, datetime):
            pd_val = pd_val.strftime("%Y-%m-%d")
        md_val = inv.get("maturity_date", "") or ""
        if isinstance(md_val, datetime):
            md_val = md_val.strftime("%Y-%m-%d")

        ws.cell(row=row, column=1, value=inv.get("name", "")).border = thin_border
        ws.cell(row=row, column=2, value=inv.get("investment_type", "")).border = thin_border
        c_inv = ws.cell(row=row, column=3, value=invested)
        c_inv.number_format = currency_fmt
        c_inv.border = thin_border
        c_cur = ws.cell(row=row, column=4, value=current)
        c_cur.number_format = currency_fmt
        c_cur.border = thin_border
        c_ret = ws.cell(row=row, column=5, value=ret)
        c_ret.number_format = currency_fmt
        c_ret.font = green_font if ret >= 0 else red_font
        c_ret.border = thin_border
        c_rp = ws.cell(row=row, column=6, value=ret_pct)
        c_rp.number_format = '0.00"%"'
        c_rp.font = green_font if ret_pct >= 0 else red_font
        c_rp.border = thin_border
        ws.cell(row=row, column=7, value=cagr).border = thin_border
        ws.cell(row=row, column=8, value=str(pd_val)).border = thin_border
        ws.cell(row=row, column=9, value=str(md_val)).border = thin_border
        ws.cell(row=row, column=10, value=inv.get("notes", "") or "").border = thin_border
        row += 1

    # Summary row
    if investments:
        total_invested = sum(i.get("invested_amount", 0) for i in investments)
        total_current = sum(i.get("current_value", 0) for i in investments)
        total_ret = total_current - total_invested
        ws.cell(row=row + 1, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=12)
        ws.cell(row=row + 1, column=3, value=total_invested).number_format = currency_fmt
        ws.cell(row=row + 1, column=4, value=total_current).number_format = currency_fmt
        ws.cell(row=row + 1, column=5, value=total_ret).number_format = currency_fmt
        ws.cell(row=row + 1, column=5).font = Font(name="Arial", bold=True, color="27AE60" if total_ret >= 0 else "C0392B")

    for col_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 35)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=investments.xlsx"}
    )


@api_router.get("/export/networth-xlsx")
async def export_networth_xlsx(request: Request):
    """Export net worth breakdown as Excel workbook"""
    user = await get_current_user(request)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    accounts = await db.accounts.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(100)
    investments = await db.investments.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(1000)
    credit_cards = await db.credit_cards.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(100)
    loans = await db.loans.find({"user_id": user.user_id, "is_active": True}, {"_id": 0}).to_list(100)
    lending = await db.lending.find({"user_id": user.user_id}, {"_id": 0}).to_list(1000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Net Worth"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    asset_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    liability_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC")
    )
    currency_fmt = '#,##0.00'

    headers = ["Category", "Item", "Type", "Amount (INR)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border

    row = 2
    # Assets
    for a in accounts:
        ws.cell(row=row, column=1, value="Asset - Account").fill = asset_fill
        ws.cell(row=row, column=2, value=a.get("name", "")).fill = asset_fill
        ws.cell(row=row, column=3, value=a.get("account_type", "")).fill = asset_fill
        c = ws.cell(row=row, column=4, value=a.get("balance", 0))
        c.number_format = currency_fmt
        c.fill = asset_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        row += 1
    for i in investments:
        ws.cell(row=row, column=1, value="Asset - Investment").fill = asset_fill
        ws.cell(row=row, column=2, value=i.get("name", "")).fill = asset_fill
        ws.cell(row=row, column=3, value=i.get("investment_type", "")).fill = asset_fill
        c = ws.cell(row=row, column=4, value=i.get("current_value", 0))
        c.number_format = currency_fmt
        c.fill = asset_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        row += 1
    for l in lending:
        if l.get("lending_type") == "lent" and not l.get("is_settled"):
            ws.cell(row=row, column=1, value="Asset - Money Lent").fill = asset_fill
            ws.cell(row=row, column=2, value=l.get("person_name", "")).fill = asset_fill
            ws.cell(row=row, column=3, value="lent").fill = asset_fill
            c = ws.cell(row=row, column=4, value=l.get("remaining_amount", l.get("amount", 0)))
            c.number_format = currency_fmt
            c.fill = asset_fill
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

    # Liabilities
    for cc in credit_cards:
        ws.cell(row=row, column=1, value="Liability - Credit Card").fill = liability_fill
        ws.cell(row=row, column=2, value=cc.get("name", "")).fill = liability_fill
        ws.cell(row=row, column=3, value="credit_card").fill = liability_fill
        c = ws.cell(row=row, column=4, value=cc.get("current_outstanding", 0))
        c.number_format = currency_fmt
        c.fill = liability_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        row += 1
    for lo in loans:
        ws.cell(row=row, column=1, value="Liability - Loan").fill = liability_fill
        ws.cell(row=row, column=2, value=lo.get("name", "")).fill = liability_fill
        ws.cell(row=row, column=3, value=lo.get("loan_type", "")).fill = liability_fill
        c = ws.cell(row=row, column=4, value=lo.get("outstanding_amount", 0))
        c.number_format = currency_fmt
        c.fill = liability_fill
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin_border
        row += 1
    for l in lending:
        if l.get("lending_type") == "borrowed" and not l.get("is_settled"):
            ws.cell(row=row, column=1, value="Liability - Borrowed").fill = liability_fill
            ws.cell(row=row, column=2, value=l.get("person_name", "")).fill = liability_fill
            ws.cell(row=row, column=3, value="borrowed").fill = liability_fill
            c = ws.cell(row=row, column=4, value=l.get("remaining_amount", l.get("amount", 0)))
            c.number_format = currency_fmt
            c.fill = liability_fill
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

    # Net Worth summary
    total_assets = sum(a.get("balance", 0) for a in accounts) + sum(i.get("current_value", 0) for i in investments) + sum(l.get("remaining_amount", 0) for l in lending if l.get("lending_type") == "lent" and not l.get("is_settled"))
    total_liabilities = sum(c.get("current_outstanding", 0) for c in credit_cards) + sum(l.get("outstanding_amount", 0) for l in loans) + sum(l.get("remaining_amount", 0) for l in lending if l.get("lending_type") == "borrowed" and not l.get("is_settled"))

    row += 1
    bold_font = Font(name="Arial", bold=True, size=12)
    ws.cell(row=row, column=1, value="Total Assets").font = bold_font
    ws.cell(row=row, column=4, value=total_assets).number_format = currency_fmt
    ws.cell(row=row, column=4).font = Font(name="Arial", bold=True, color="27AE60", size=12)
    row += 1
    ws.cell(row=row, column=1, value="Total Liabilities").font = bold_font
    ws.cell(row=row, column=4, value=total_liabilities).number_format = currency_fmt
    ws.cell(row=row, column=4).font = Font(name="Arial", bold=True, color="C0392B", size=12)
    row += 1
    ws.cell(row=row, column=1, value="NET WORTH").font = Font(name="Arial", bold=True, size=14)
    nw = total_assets - total_liabilities
    ws.cell(row=row, column=4, value=nw).number_format = currency_fmt
    ws.cell(row=row, column=4).font = Font(name="Arial", bold=True, color="1B4F72", size=14)

    for col_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 35)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=networth.xlsx"}
    )


# ==================== CLOUD DRIVE INTEGRATION ====================

class CloudDriveConnect(BaseModel):
    provider: str  # google_drive, onedrive, dropbox
    access_token: str
    refresh_token: Optional[str] = None
    token_expiry: Optional[str] = None

class CloudDriveUploadRequest(BaseModel):
    provider: str  # google_drive, onedrive, dropbox
    export_type: str  # transactions, investments, networth, full_backup
    file_format: str = "xlsx"  # xlsx, csv, json
    folder_path: Optional[str] = None  # target folder in cloud drive

@api_router.post("/cloud-drive/connect")
async def connect_cloud_drive(data: CloudDriveConnect, request: Request):
    """Store cloud drive OAuth credentials for the user"""
    user = await get_current_user(request)

    drive_doc = {
        "user_id": user.user_id,
        "provider": data.provider,
        "access_token": data.access_token,
        "refresh_token": data.refresh_token,
        "token_expiry": data.token_expiry,
        "connected_at": datetime.now(timezone.utc),
        "is_active": True
    }

    # Upsert — one connection per provider per user
    await db.cloud_drives.update_one(
        {"user_id": user.user_id, "provider": data.provider},
        {"$set": drive_doc},
        upsert=True
    )

    # Update user settings to reflect preferred storage provider
    await db.user_settings.update_one(
        {"user_id": user.user_id},
        {"$set": {"storage_provider": data.provider, "updated_at": datetime.now(timezone.utc)}},
        upsert=True
    )

    return {"message": f"{data.provider} connected successfully", "provider": data.provider}


@api_router.get("/cloud-drive/status")
async def cloud_drive_status(request: Request):
    """Get connected cloud drive status"""
    user = await get_current_user(request)
    drives = await db.cloud_drives.find(
        {"user_id": user.user_id, "is_active": True}, {"_id": 0, "access_token": 0, "refresh_token": 0}
    ).to_list(10)
    return {"connected_drives": drives}


@api_router.delete("/cloud-drive/{provider}")
async def disconnect_cloud_drive(provider: str, request: Request):
    """Disconnect a cloud drive"""
    user = await get_current_user(request)
    result = await db.cloud_drives.update_one(
        {"user_id": user.user_id, "provider": provider},
        {"$set": {"is_active": False}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Drive connection not found")
    return {"message": f"{provider} disconnected"}


@api_router.post("/cloud-drive/upload")
async def upload_to_cloud_drive(data: CloudDriveUploadRequest, request: Request):
    """Export data and upload to connected cloud drive (Google Drive / OneDrive / Dropbox)"""
    user = await get_current_user(request)

    # 1. Verify cloud drive is connected
    drive = await db.cloud_drives.find_one(
        {"user_id": user.user_id, "provider": data.provider, "is_active": True}
    )
    if not drive:
        raise HTTPException(status_code=400, detail=f"{data.provider} is not connected. Connect it first via /api/cloud-drive/connect")

    access_token = drive["access_token"]
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    # 2. Generate the export file
    if data.file_format == "xlsx":
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if data.export_type == "transactions":
            buf = await _build_transactions_xlsx(user.user_id)
            filename = f"BillTracker_Transactions_{timestamp}.xlsx"
        elif data.export_type == "investments":
            # Re-use the investments xlsx builder inline
            response = await export_investments_xlsx(request)
            file_bytes = response.body
            filename = f"BillTracker_Investments_{timestamp}.xlsx"
        elif data.export_type == "networth":
            response = await export_networth_xlsx(request)
            file_bytes = response.body
            filename = f"BillTracker_NetWorth_{timestamp}.xlsx"
        else:
            raise HTTPException(status_code=400, detail="Invalid export_type for xlsx")
        file_bytes = buf.getvalue() if data.export_type == "transactions" else file_bytes
    elif data.file_format == "csv":
        content_type = "text/csv"
        if data.export_type == "transactions":
            response = await export_transactions_csv(request)
            file_bytes = response.body
            filename = f"BillTracker_Transactions_{timestamp}.csv"
        elif data.export_type == "investments":
            response = await export_investments_csv(request)
            file_bytes = response.body
            filename = f"BillTracker_Investments_{timestamp}.csv"
        elif data.export_type == "networth":
            response = await export_networth_csv(request)
            file_bytes = response.body
            filename = f"BillTracker_NetWorth_{timestamp}.csv"
        else:
            raise HTTPException(status_code=400, detail="Invalid export_type for csv")
    elif data.file_format == "json":
        content_type = "application/json"
        response_data = await export_data(request)
        file_bytes = json.dumps(response_data, default=str, indent=2).encode('utf-8')
        filename = f"BillTracker_FullBackup_{timestamp}.json"
    else:
        raise HTTPException(status_code=400, detail="Unsupported file_format. Use: xlsx, csv, json")

    # 3. Upload to the cloud drive
    upload_result = None
    try:
        async with httpx.AsyncClient(timeout=30.0) as http_client:
            if data.provider == "google_drive":
                # Google Drive Files API — simple upload
                metadata = {"name": filename}
                if data.folder_path:
                    metadata["parents"] = [data.folder_path]

                # Step 1: Create metadata
                meta_resp = await http_client.post(
                    "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart",
                    headers={"Authorization": f"Bearer {access_token}"},
                    files={
                        "metadata": ("metadata", json.dumps(metadata).encode(), "application/json"),
                        "file": (filename, file_bytes, content_type)
                    }
                )
                if meta_resp.status_code in (200, 201):
                    upload_result = meta_resp.json()
                else:
                    raise HTTPException(status_code=502, detail=f"Google Drive upload failed: {meta_resp.text}")

            elif data.provider == "onedrive":
                # OneDrive Files API — simple upload
                folder = data.folder_path or "BillTracker"
                upload_url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{folder}/{filename}:/content"
                od_resp = await http_client.put(
                    upload_url,
                    headers={
                        "Authorization": f"Bearer {access_token}",
                        "Content-Type": content_type
                    },
                    content=file_bytes
                )
                if od_resp.status_code in (200, 201):
                    upload_result = od_resp.json()
                else:
                    raise HTTPException(status_code=502, detail=f"OneDrive upload failed: {od_resp.text}")

            elif data.provider == "dropbox":
                # Dropbox Files API
                folder = data.folder_path or "/BillTracker"
                dropbox_path = f"{folder}/{filename}"
                db_resp = await http_client.post(
                    "https://content.dropboxapi.com/2/files/upload",
                    headers={
                        "Authorization": f"Bearer {access_token}",
                        "Content-Type": "application/octet-stream",
                        "Dropbox-API-Arg": json.dumps({"path": dropbox_path, "mode": "add", "autorename": True})
                    },
                    content=file_bytes
                )
                if db_resp.status_code == 200:
                    upload_result = db_resp.json()
                else:
                    raise HTTPException(status_code=502, detail=f"Dropbox upload failed: {db_resp.text}")

            else:
                raise HTTPException(status_code=400, detail=f"Unsupported provider: {data.provider}")

    except httpx.RequestError as e:
        raise HTTPException(status_code=502, detail=f"Network error uploading to {data.provider}: {str(e)}")

    # 4. Log the upload
    await db.export_logs.insert_one({
        "user_id": user.user_id,
        "provider": data.provider,
        "export_type": data.export_type,
        "file_format": data.file_format,
        "filename": filename,
        "upload_result": upload_result,
        "created_at": datetime.now(timezone.utc)
    })

    return {
        "message": f"File uploaded to {data.provider} successfully",
        "filename": filename,
        "provider": data.provider,
        "details": upload_result
    }


@api_router.get("/export/history")
async def get_export_history(request: Request):
    """Get history of cloud drive exports"""
    user = await get_current_user(request)
    logs = await db.export_logs.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    return {"exports": logs}


# ==================== DATA BACKUP & RESTORE ====================

@api_router.get("/backup")
async def create_backup(request: Request):
    """Create a full data backup as downloadable JSON"""
    user = await get_current_user(request)

    backup = {
        "backup_version": "2.0",
        "app": "Bill Tracker",
        "user_id": user.user_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "data": {}
    }

    collections = [
        ("accounts", {"user_id": user.user_id}),
        ("income", {"user_id": user.user_id}),
        ("expenses", {"user_id": user.user_id}),
        ("bills", {"user_id": user.user_id}),
        ("payments", {"user_id": user.user_id}),
        ("categories", {"user_id": user.user_id}),
        ("budgets", {"user_id": user.user_id}),
        ("family_members", {"user_id": user.user_id}),
        ("credit_cards", {"user_id": user.user_id}),
        ("loans", {"user_id": user.user_id}),
        ("lending", {"user_id": user.user_id}),
        ("investments", {"user_id": user.user_id}),
        ("reminders", {"user_id": user.user_id}),
        ("rentals", {"user_id": user.user_id}),
        ("investment_headings", {"user_id": user.user_id}),
        ("user_settings", {"user_id": user.user_id}),
    ]

    for coll_name, query in collections:
        docs = await db[coll_name].find(query, {"_id": 0}).to_list(50000)
        backup["data"][coll_name] = docs

    backup_json = json.dumps(backup, default=str, indent=2)
    return Response(
        content=backup_json,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=billtracker_backup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"}
    )


@api_router.post("/backup/restore")
async def restore_backup(request: Request):
    """Restore data from a JSON backup file"""
    user = await get_current_user(request)
    body = await request.json()

    if body.get("backup_version") not in ("1.0", "2.0"):
        raise HTTPException(status_code=400, detail="Invalid backup file format")

    backup_data = body.get("data", {})
    if not backup_data:
        raise HTTPException(status_code=400, detail="No data found in backup")

    restored = {}
    for coll_name, docs in backup_data.items():
        if not isinstance(docs, list):
            continue
        # Delete existing data for this user in this collection
        await db[coll_name].delete_many({"user_id": user.user_id})
        # Insert backup data
        if docs:
            # Ensure all docs have the current user_id
            for doc in docs:
                doc["user_id"] = user.user_id
            await db[coll_name].insert_many(docs)
        restored[coll_name] = len(docs)

    return {
        "message": "Backup restored successfully",
        "collections_restored": restored,
        "restored_at": datetime.now(timezone.utc).isoformat()
    }


# ==================== NOTES ENDPOINTS ====================

@api_router.post("/notes")
async def create_note(data: NoteCreate, request: Request):
    user = await get_current_user(request)
    note_id = f"note_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    note = {
        "note_id": note_id, "user_id": user.user_id,
        "title": data.title, "content": data.content,
        "sections": data.sections or [],
        "tags": data.tags or [],
        "linked_type": data.linked_type, "linked_id": data.linked_id,
        "priority": data.priority or "normal",
        "color": data.color,
        "is_archived": False,
        "created_at": now, "updated_at": now,
    }
    await db.notes.insert_one(note)
    note.pop("_id", None)
    return note

@api_router.get("/notes")
async def get_notes(request: Request, tag: Optional[str] = None, linked_type: Optional[str] = None, is_archived: bool = False):
    user = await get_current_user(request)
    query = {"user_id": user.user_id, "is_archived": is_archived}
    if tag:
        query["tags"] = tag
    if linked_type:
        query["linked_type"] = linked_type
    notes = await db.notes.find(query, {"_id": 0}).sort("updated_at", -1).to_list(200)
    return notes

@api_router.get("/notes/{note_id}")
async def get_note(note_id: str, request: Request):
    user = await get_current_user(request)
    note = await db.notes.find_one({"note_id": note_id, "user_id": user.user_id}, {"_id": 0})
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    return note

@api_router.put("/notes/{note_id}")
async def update_note(note_id: str, data: NoteUpdate, request: Request):
    user = await get_current_user(request)
    existing = await db.notes.find_one({"note_id": note_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Note not found")
    update_data = {k: v for k, v in data.model_dump(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    await db.notes.update_one({"note_id": note_id}, {"$set": update_data})
    updated = await db.notes.find_one({"note_id": note_id}, {"_id": 0})
    return updated

@api_router.delete("/notes/{note_id}")
async def delete_note(note_id: str, request: Request):
    user = await get_current_user(request)
    existing = await db.notes.find_one({"note_id": note_id, "user_id": user.user_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Note not found")
    await db.notes.delete_one({"note_id": note_id})
    return {"message": "Note deleted"}


# ==================== PORTFOLIO ANALYTICS ENDPOINT ====================

@api_router.get("/portfolio/analytics")
async def get_portfolio_analytics(request: Request):
    """Get investment portfolio analytics with ROI and CAGR calculations"""
    user = await get_current_user(request)
    investments = await db.investments.find({"user_id": user.user_id}, {"_id": 0}).to_list(500)

    total_invested = sum(i.get("invested_amount", 0) for i in investments)
    total_current = sum(i.get("current_value", 0) for i in investments)
    total_gain = total_current - total_invested
    roi_pct = (total_gain / total_invested * 100) if total_invested > 0 else 0

    # Calculate CAGR for each investment
    now = datetime.now(timezone.utc)
    enriched = []
    for inv in investments:
        invested = inv.get("invested_amount", 0)
        current = inv.get("current_value", 0)
        gain = current - invested
        inv_roi = (gain / invested * 100) if invested > 0 else 0

        cagr = 0
        pd_str = inv.get("purchase_date", "")
        if pd_str and invested > 0 and current > 0:
            try:
                pd = datetime.fromisoformat(pd_str.replace("Z", "+00:00"))
                years = max((now - pd).days / 365.25, 0.01)
                cagr = ((current / invested) ** (1 / years) - 1) * 100
            except Exception:
                pass

        enriched.append({
            **inv,
            "gain": round(gain, 2),
            "roi_pct": round(inv_roi, 2),
            "cagr_pct": round(cagr, 2),
        })

    # By type breakdown
    by_type = {}
    for inv in enriched:
        t = inv.get("investment_type", "other")
        if t not in by_type:
            by_type[t] = {"invested": 0, "current": 0, "count": 0}
        by_type[t]["invested"] += inv.get("invested_amount", 0)
        by_type[t]["current"] += inv.get("current_value", 0)
        by_type[t]["count"] += 1

    type_breakdown = []
    for t, v in by_type.items():
        g = v["current"] - v["invested"]
        type_breakdown.append({
            "type": t,
            "invested": round(v["invested"], 2),
            "current": round(v["current"], 2),
            "gain": round(g, 2),
            "roi_pct": round((g / v["invested"] * 100) if v["invested"] > 0 else 0, 2),
            "count": v["count"],
            "allocation_pct": round((v["current"] / total_current * 100) if total_current > 0 else 0, 2),
        })

    return {
        "total_invested": round(total_invested, 2),
        "total_current": round(total_current, 2),
        "total_gain": round(total_gain, 2),
        "overall_roi_pct": round(roi_pct, 2),
        "investment_count": len(investments),
        "type_breakdown": type_breakdown,
        "investments": enriched,
    }


# ==================== MPIN ENDPOINTS ====================

@api_router.post("/mpin/setup")
async def setup_mpin(request: Request):
    """Set up or update MPIN for quick app access"""
    user = await get_current_user(request)
    body = await request.json()
    mpin = body.get("mpin", "")

    if not mpin or len(mpin) < 4 or len(mpin) > 6 or not mpin.isdigit():
        raise HTTPException(status_code=400, detail="MPIN must be 4-6 digits")

    hashed = bcrypt.hashpw(mpin.encode(), bcrypt.gensalt()).decode()
    now = datetime.now(timezone.utc)

    await db.user_mpin.update_one(
        {"user_id": user.user_id},
        {"$set": {"mpin_hash": hashed, "is_enabled": True, "updated_at": now}},
        upsert=True
    )
    return {"message": "MPIN set successfully", "is_enabled": True}

@api_router.post("/mpin/verify")
async def verify_mpin(request: Request):
    """Verify MPIN for app unlock"""
    user = await get_current_user(request)
    body = await request.json()
    mpin = body.get("mpin", "")

    record = await db.user_mpin.find_one({"user_id": user.user_id})
    if not record or not record.get("is_enabled"):
        raise HTTPException(status_code=404, detail="MPIN not set up")

    if bcrypt.checkpw(mpin.encode(), record["mpin_hash"].encode()):
        return {"verified": True}
    raise HTTPException(status_code=401, detail="Invalid MPIN")

@api_router.get("/mpin/status")
async def mpin_status(request: Request):
    """Check if MPIN is enabled for user"""
    user = await get_current_user(request)
    record = await db.user_mpin.find_one({"user_id": user.user_id})
    return {"is_enabled": bool(record and record.get("is_enabled", False))}

@api_router.post("/mpin/disable")
async def disable_mpin(request: Request):
    """Disable MPIN"""
    user = await get_current_user(request)
    await db.user_mpin.update_one(
        {"user_id": user.user_id},
        {"$set": {"is_enabled": False, "updated_at": datetime.now(timezone.utc)}}
    )
    return {"message": "MPIN disabled", "is_enabled": False}


# ==================== CALENDAR ENDPOINTS ====================

@api_router.get("/calendar/events")
async def get_calendar_events(request: Request, month: int = None, year: int = None):
    """Get all financial events for a given month for calendar view"""
    user = await get_current_user(request)

    now = datetime.now(timezone.utc)
    m = month or now.month
    y = year or now.year

    start = datetime(y, m, 1, tzinfo=timezone.utc)
    if m == 12:
        end = datetime(y + 1, 1, 1, tzinfo=timezone.utc)
    else:
        end = datetime(y, m + 1, 1, tzinfo=timezone.utc)

    events = []

    # Bills
    bills = await db.bills.find({
        "user_id": user.user_id,
        "due_date": {"$gte": start.isoformat(), "$lt": end.isoformat()}
    }, {"_id": 0}).to_list(200)
    for b in bills:
        events.append({
            "id": b.get("bill_id"),
            "date": b.get("due_date", "")[:10],
            "title": b.get("name", "Bill"),
            "type": "bill",
            "amount": b.get("amount", 0),
            "status": b.get("payment_status", "unpaid"),
            "color": "#EF4444",
        })

    # Income
    incomes = await db.income.find({
        "user_id": user.user_id,
        "date": {"$gte": start.isoformat(), "$lt": end.isoformat()}
    }, {"_id": 0}).to_list(200)
    for i in incomes:
        events.append({
            "id": i.get("income_id"),
            "date": i.get("date", "")[:10],
            "title": i.get("source", "Income"),
            "type": "income",
            "amount": i.get("amount", 0),
            "color": "#22C55E",
        })

    # Expenses
    expenses = await db.expenses.find({
        "user_id": user.user_id,
        "date": {"$gte": start.isoformat(), "$lt": end.isoformat()}
    }, {"_id": 0}).to_list(200)
    for e in expenses:
        events.append({
            "id": e.get("expense_id"),
            "date": e.get("date", "")[:10],
            "title": e.get("description", e.get("category", "Expense")),
            "type": "expense",
            "amount": e.get("amount", 0),
            "color": "#F59E0B",
        })

    # Reminders
    reminders = await db.reminders.find({
        "user_id": user.user_id,
        "due_date": {"$gte": start.isoformat(), "$lt": end.isoformat()}
    }, {"_id": 0}).to_list(200)
    for r in reminders:
        events.append({
            "id": r.get("reminder_id"),
            "date": r.get("due_date", "")[:10],
            "title": r.get("title", "Reminder"),
            "type": "reminder",
            "amount": r.get("amount", 0),
            "color": "#8B5CF6",
        })

    return {"month": m, "year": y, "events": events}


# Include router and add middleware
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)
